from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
import subprocess
import sys
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
import zipfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PYTHON = Path(sys.executable)
sys.path.insert(0, str(ROOT / "src"))

from screen_mouse_recorder.app import ScreenMouseRecorderApp
from screen_mouse_recorder.analysis import generate_behavior_report
from screen_mouse_recorder.calibration import build_calibration_result
from screen_mouse_recorder.config import AppConfig
from screen_mouse_recorder.diagnostics.error_report import (
    build_error_report,
    format_error_dialog_message,
    write_error_report,
)
from screen_mouse_recorder.diagnostics.service import ErrorReporter
from screen_mouse_recorder.event_extraction import (
    OCREventExtractionConfig,
    OCRTextItem,
    extract_selected_ocr_events,
)
from screen_mouse_recorder.frame_sampler import (
    ClickKeyframeConfig,
    ClickMarker,
    CropRegion,
    DenseRange,
    FrameSamplerConfig,
    FramePlanEntry,
    VideoInfo,
    _add_silent_gap_keyframes,
    _compose_sheet,
    _nearest_click_marker,
    _prepare_frame_image,
    build_click_keyframe_plan,
    build_frame_plan,
    default_output_dir,
    estimate_click_keyframe_sampling,
    estimate_sampling,
    format_timecode,
    load_click_keyframe_events,
    load_click_markers,
    parse_timecode,
    run_frame_export,
    select_click_keyframes,
    select_click_keyframes_with_stats,
)
from screen_mouse_recorder.frame_export.ui_state import (
    ClickKeyframeFormState,
    FrameSamplerFormState,
    build_click_keyframe_config_from_state,
    build_frame_sampler_config_from_state,
    collect_dense_ranges,
)
from screen_mouse_recorder.frame_export.click_keyframes import build_click_keyframe_visual_signatures
from screen_mouse_recorder.frame_export.progress import (
    completed_progress,
    failed_progress,
    format_duration_seconds,
    starting_progress,
    update_progress,
)
from PIL import Image
from screen_mouse_recorder.models import Region, TimingContext
from screen_mouse_recorder.mouse_logger import MouseActivityLogger
from screen_mouse_recorder.naming import build_session_id, default_report_output_dir, sanitize_session_name
from screen_mouse_recorder.postprocess import generate_summary
from screen_mouse_recorder.reporting.service import make_behavior_report_job
from screen_mouse_recorder.storage import JsonlWriter, SessionStorage
from screen_mouse_recorder.updater import check_for_updates
from screen_mouse_recorder.video_recorder import FFmpegRecorder


class CoreSmokeTests(unittest.TestCase):
    def test_manual_ocr_event_output_keeps_index_time_and_review_links(self) -> None:
        class FakeOCREngine:
            name = "fake-ocr"
            version = "1.0"

            def recognize(self, image_path: Path) -> tuple[list[OCRTextItem], float]:
                if "000001" in image_path.name:
                    return (
                        [
                            OCRTextItem("#022", 0.99, (5, 5, 50, 25)),
                            OCRTextItem("00:01:45", 0.98, (5, 28, 90, 48)),
                            OCRTextItem("单人BOSS", 0.97, (90, 180, 210, 215)),
                            OCRTextItem("新功能开启", 0.99, (70, 230, 230, 270)),
                            OCRTextItem("任务奖励", 0.96, (110, 320, 190, 345)),
                        ],
                        0.05,
                    )
                return (
                    [
                        OCRTextItem("#147", 0.99, (5, 5, 50, 25)),
                        OCRTextItem("00:10:10", 0.98, (5, 28, 90, 48)),
                        OCRTextItem("新技能解锁", 0.99, (70, 120, 230, 160)),
                        OCRTextItem("大圣归来", 0.97, (100, 180, 200, 215)),
                        OCRTextItem("技能效果", 0.96, (110, 250, 190, 275)),
                    ],
                    0.06,
                )

        with TemporaryDirectory() as directory:
            root = Path(directory)
            source_one = root / "feature.png"
            source_two = root / "skill.png"
            Image.new("RGB", (300, 500), "#202020").save(source_one)
            Image.new("RGB", (300, 500), "#303030").save(source_two)
            index_path = root / "keyframes_click_sheet_index.json"
            index_path.write_text(
                json.dumps(
                    {
                        "frames": [
                            {
                                "index": 22,
                                "seconds": 105.0,
                                "timestamp": "00:01:45",
                                "sheet": "keyframes_click_sheet.png",
                                "sheet_row": 5,
                                "sheet_col": 2,
                            },
                            {
                                "index": 147,
                                "seconds": 610.0,
                                "timestamp": "00:10:10",
                                "sheet": "keyframes_click_sheet.png",
                                "sheet_row": 5,
                                "sheet_col": 3,
                            },
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            selection_path = root / "selected_ocr_tiles.json"
            selection_path.write_text(
                json.dumps(
                    {
                        "schema_version": "1.0",
                        "source_index": str(index_path),
                        "selections": [
                            {"tile_index": 22, "source_frame": str(source_one)},
                            {"tile_index": 147, "source_frame": str(source_two)},
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            result = extract_selected_ocr_events(
                OCREventExtractionConfig(
                    index_json=index_path,
                    selection_json=selection_path,
                    output_dir=root / "ocr_events",
                ),
                engine=FakeOCREngine(),
            )

            payload = json.loads(result.json_path.read_text(encoding="utf-8"))
            selection_payload = json.loads(selection_path.read_text(encoding="utf-8"))
            selection_schema = json.loads(
                (ROOT / "schemas" / "selected_ocr_tiles.schema.json").read_text(encoding="utf-8")
            )
            result_schema = json.loads(
                (ROOT / "schemas" / "event_ocr_results.schema.json").read_text(encoding="utf-8")
            )
            self.assertEqual(result.event_count, 2)
            self.assertEqual(result.needs_review_count, 0)
            self.assertTrue(set(selection_schema["required"]).issubset(selection_payload))
            self.assertTrue(set(result_schema["required"]).issubset(payload))
            self.assertTrue(set(result_schema["properties"]["source"]["required"]).issubset(payload["source"]))
            self.assertTrue(set(result_schema["$defs"]["event"]["required"]).issubset(payload["events"][0]))
            self.assertEqual(payload["source"]["ocr_engine"], "fake-ocr")
            self.assertEqual(payload["events"][0]["event_type"], "new_feature_unlocked")
            self.assertEqual(payload["events"][0]["event_name"], "单人BOSS")
            self.assertEqual(payload["events"][0]["timestamp"], "00:01:45")
            self.assertEqual(payload["events"][0]["time_source"], "index_json")
            self.assertEqual(payload["events"][0]["time_check"], "matched")
            self.assertEqual(payload["events"][1]["event_type"], "new_skill_unlocked")
            self.assertEqual(payload["events"][1]["event_name"], "大圣归来")
            self.assertTrue(Path(payload["events"][0]["source_frame"]).exists())
            self.assertTrue(Path(payload["events"][0]["review_image"]).exists())
            workbook = load_workbook(result.xlsx_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.active.max_row, 3)
            finally:
                workbook.close()

    def test_error_report_has_ocr_codes(self) -> None:
        self.assertEqual(build_error_report("ocr_input", ValueError("bad selection")).code, "OCR-INPUT-001")
        self.assertEqual(build_error_report("ocr_run", RuntimeError("model failed")).code, "OCR-RUN-001")
        self.assertEqual(
            build_error_report("ocr_region_scan", RuntimeError("profile mismatch")).code,
            "OCR-REGION-SCAN-001",
        )

    def test_update_check_ignores_non_git_folder(self) -> None:
        with TemporaryDirectory() as directory:
            status = check_for_updates(Path(directory))

        self.assertFalse(status.available)
        self.assertIsNone(status.git_root)

    def test_error_report_writes_code_explanation_and_context(self) -> None:
        with TemporaryDirectory() as directory:
            report = build_error_report(
                "frame_export",
                RuntimeError("disk full"),
                context={"output_dir": Path(directory) / "frame_exports"},
            )
            json_path, txt_path = write_error_report(Path(directory), report)

            dialog = format_error_dialog_message(report, txt_path)

            self.assertEqual(report.code, "FRM-EXPORT-001")
            self.assertTrue(json_path.exists())
            self.assertTrue(txt_path.exists())
            self.assertIn("\u9519\u8bef\u4ee3\u7801\uff1aFRM-EXPORT-001", dialog)
            self.assertIn("\u62bd\u5e27\u62fc\u56fe\u5931\u8d25", txt_path.read_text(encoding="utf-8"))
            self.assertIn("output_dir", json_path.read_text(encoding="utf-8"))

    def test_error_report_has_summary_regenerate_code(self) -> None:
        report = build_error_report("summary_regenerate", RuntimeError("bad rows"))

        self.assertEqual(report.code, "SUM-REGEN-001")
        self.assertIn("\u6458\u8981\u8868\u683c", report.title)

    def test_error_reporter_service_writes_report_files(self) -> None:
        with TemporaryDirectory() as directory:
            result = ErrorReporter(Path(directory)).create(
                "frame_export",
                RuntimeError("disk full"),
                {"output_dir": Path(directory) / "frame_exports"},
            )

            self.assertEqual(result.report.code, "FRM-EXPORT-001")
            self.assertIsNotNone(result.json_path)
            self.assertIsNotNone(result.txt_path)
            self.assertTrue(result.json_path.exists() if result.json_path else False)
            self.assertTrue(result.txt_path.exists() if result.txt_path else False)

    def test_region_mapping(self) -> None:
        region = Region(screen_x=100, screen_y=200, width=400, height=300)
        mapped = region.map_point(300, 350)
        self.assertEqual(mapped["region_x"], 200)
        self.assertEqual(mapped["region_y"], 150)
        self.assertEqual(mapped["region_x_norm"], 0.5)
        self.assertEqual(mapped["region_y_norm"], 0.5)
        self.assertTrue(mapped["inside_region"])

    def test_region_even_sized_for_h264(self) -> None:
        region = Region(screen_x=-1544, screen_y=70, width=951, height=783)
        normalized = region.even_sized()

        self.assertEqual(normalized.width, 950)
        self.assertEqual(normalized.height, 782)
        self.assertEqual(normalized.screen_x, region.screen_x)
        self.assertEqual(normalized.screen_y, region.screen_y)

    def test_paused_duration_is_removed_from_video_time(self) -> None:
        timing = TimingContext(
            session_id="test",
            logger_start_monotonic_ms=1000,
            video_zero_monotonic_ms=1200,
            paused_duration_ms=500,
        )
        self.assertEqual(timing.t_video_ms(2200), 500)
        self.assertEqual(timing.timing_dict()["paused_duration_ms"], 500)

    def test_session_name_sanitizer(self) -> None:
        self.assertEqual(ScreenMouseRecorderApp._sanitize_session_name(" demo / round:1 "), "demo_round_1")
        self.assertEqual(ScreenMouseRecorderApp._sanitize_session_name("中文 任务"), "中文_任务")
        self.assertEqual(sanitize_session_name(" demo / round:1 "), "demo_round_1")

    def test_build_session_id_uses_rec_prefix_and_safe_suffix(self) -> None:
        session_id = build_session_id("中文 任务", now=datetime(2026, 7, 7, 20, 30, 5))

        self.assertEqual(session_id, "rec_20260707_203005_中文_任务")

    def test_default_report_output_dir_prefers_new_name_but_reads_legacy(self) -> None:
        with TemporaryDirectory() as directory:
            session_dir = Path(directory) / "rec_20260707_203005"
            session_dir.mkdir()
            self.assertEqual(default_report_output_dir(session_dir), session_dir / "auto_report")

            legacy_dir = session_dir / "analysis_output"
            legacy_dir.mkdir()
            self.assertEqual(default_report_output_dir(session_dir), legacy_dir)

    def test_behavior_report_job_uses_default_output_and_ffmpeg_path(self) -> None:
        with TemporaryDirectory() as directory:
            session_dir = Path(directory) / "rec_20260707_203005"
            session_dir.mkdir()

            job = make_behavior_report_job(session_dir, ffmpeg_path="ffmpeg-test")

        self.assertEqual(job.source_path, session_dir.resolve())
        self.assertEqual(job.output_dir, (session_dir / "auto_report").resolve())
        self.assertEqual(job.ffmpeg_path, "ffmpeg-test")

    def test_config_includes_recording_status_banner_toggle(self) -> None:
        config = AppConfig(show_recording_status_banner=False)

        self.assertFalse(config.show_recording_status_banner)
        self.assertFalse(config.to_dict()["show_recording_status_banner"])

    def test_config_includes_frame_sampler_defaults(self) -> None:
        config = AppConfig()

        self.assertEqual(config.frame_sampler_interval_seconds, 10.0)
        self.assertEqual(config.frame_sampler_cols, 5)
        self.assertEqual(config.frame_sampler_rows, 6)
        self.assertTrue(config.frame_sampler_show_timestamp)
        self.assertEqual(config.frame_sampler_mode, "interval")
        self.assertEqual(config.frame_sampler_output_root, "frame_exports")
        self.assertEqual(config.frame_sampler_keyframe_max_frames, 0)
        self.assertEqual(config.frame_sampler_keyframe_time_dedupe_ms, 1500)
        self.assertEqual(config.frame_sampler_keyframe_distance_dedupe_px, 80)
        self.assertEqual(config.frame_sampler_keyframe_visual_threshold_percent, 22)
        self.assertTrue(config.frame_sampler_draw_click_markers)

    def test_frame_sampler_timecode_helpers(self) -> None:
        self.assertEqual(parse_timecode("01:02:03"), 3723)
        self.assertEqual(parse_timecode("02:30"), 150)
        self.assertEqual(format_timecode(62.5), "00:01:02.500")

    def test_frame_sampler_ui_state_builds_interval_config(self) -> None:
        config = build_frame_sampler_config_from_state(
            FrameSamplerFormState(
                video_path=Path("video.mp4"),
                output_dir=Path("out"),
                output_name="我的/导出:01",
                start_text="00:01",
                end_text="00:10",
                interval_text="2.5",
                cols_text="20",
                rows_text="bad",
                thumb_width_text="80",
                quality_preset="\u65e0\u635f",
                crop_enabled=True,
                crop_x_text="10",
                crop_y_text="20",
                crop_width_text="300",
                crop_height_text="120",
                dense_rows=[{"start": "00:03", "end": "00:05", "interval": "0.05"}],
                click_match_window_seconds=0.05,
            )
        )

        self.assertEqual(config.start_seconds, 1)
        self.assertEqual(config.end_seconds, 10)
        self.assertEqual(config.sheet_cols, 12)
        self.assertEqual(config.sheet_rows, 6)
        self.assertEqual(config.thumb_width, 120)
        self.assertEqual(config.output_format, "png")
        self.assertEqual(config.output_basename, "我的_导出_01")
        self.assertEqual(config.crop, CropRegion(10, 20, 300, 120))
        self.assertEqual(config.dense_ranges[0], DenseRange(3, 5, 0.1))
        self.assertEqual(config.click_match_window_seconds, 0.1)

    def test_frame_sampler_ui_state_builds_click_keyframe_config(self) -> None:
        config = build_click_keyframe_config_from_state(
            ClickKeyframeFormState(
                video_path=Path("video.mp4"),
                events_path=Path("mouse_events.jsonl"),
                output_dir=Path("out"),
                output_name="点击版*01",
                max_frames_text="150",
                time_dedupe_ms_text="1200",
                distance_dedupe_px_text="60",
                visual_threshold_percent_text="35",
            )
        )

        self.assertEqual(config.max_frames, 150)
        self.assertEqual(config.time_dedupe_seconds, 1.2)
        self.assertEqual(config.distance_dedupe_px, 60)
        self.assertEqual(config.visual_change_threshold, 0.35)
        self.assertEqual(config.output_basename, "点击版_01")

    def test_frame_sampler_ui_state_rejects_incomplete_dense_range(self) -> None:
        with self.assertRaises(ValueError):
            collect_dense_ranges([{"start": "00:03", "end": "", "interval": "2"}])

    def test_frame_export_progress_formats_eta(self) -> None:
        progress = update_progress(25, 100, "抽帧", started_ms=0, now_ms=10_000)

        self.assertEqual(progress.percent, 25)
        self.assertEqual(progress.progress_text, "25/100 · 抽帧")
        self.assertEqual(progress.remaining_text, "预计剩余 30秒")

    def test_frame_export_progress_handles_edge_states(self) -> None:
        self.assertEqual(starting_progress().remaining_text, "预计剩余 --")
        self.assertEqual(failed_progress().remaining_text, "生成失败")
        self.assertEqual(completed_progress().percent, 100)
        self.assertEqual(update_progress(0, 0, "准备", started_ms=None, now_ms=0).progress_text, "准备")
        self.assertEqual(format_duration_seconds(3661), "1小时01分")

    def test_frame_sampler_default_output_dir_uses_parent_folder_and_range(self) -> None:
        path = default_output_dir(
            Path(r"D:\sessions\青云决\recording.mp4"),
            Path(r"D:\sessions\青云决"),
            start_seconds=90,
            end_seconds=1800,
        )

        self.assertEqual(path.name, "interval_000130-003000_full_v001")

    def test_frame_sampler_default_output_dir_uses_next_available_index(self) -> None:
        with TemporaryDirectory() as directory:
            folder = Path(directory) / "青云决"
            folder.mkdir()
            (folder / "interval_000130-003000_full_v001").mkdir()

            path = default_output_dir(
                folder / "recording.mp4",
                folder,
                start_seconds=90,
                end_seconds=1800,
            )

        self.assertEqual(path.name, "interval_000130-003000_full_v002")

    def test_frame_sampler_plan_uses_dense_range_and_dedupes(self) -> None:
        video = VideoInfo(Path("demo.mp4"), duration_seconds=40, width=100, height=100, fps=30, file_size_bytes=1000)
        config = FrameSamplerConfig(
            video_path=Path("demo.mp4"),
            output_dir=Path("out"),
            start_seconds=0,
            end_seconds=30,
            interval_seconds=10,
            sheet_cols=3,
            sheet_rows=2,
            dense_start_seconds=5,
            dense_end_seconds=15,
            dense_interval_seconds=5,
        )

        plan = build_frame_plan(config, video)
        self.assertEqual([entry.seconds for entry in plan], [0, 5, 10, 15, 20, 30])
        self.assertEqual([entry.is_dense for entry in plan], [False, True, True, True, False, False])
        self.assertEqual(plan[-1].sheet_row, 2)
        self.assertEqual(plan[-1].sheet_col, 3)

    def test_frame_sampler_plan_uses_multiple_dense_ranges(self) -> None:
        video = VideoInfo(Path("demo.mp4"), duration_seconds=60, width=100, height=100, fps=30, file_size_bytes=1000)
        config = FrameSamplerConfig(
            video_path=Path("demo.mp4"),
            output_dir=Path("out"),
            start_seconds=0,
            end_seconds=30,
            interval_seconds=10,
            sheet_cols=4,
            sheet_rows=2,
            dense_ranges=[
                DenseRange(start_seconds=4, end_seconds=8, interval_seconds=2),
                DenseRange(start_seconds=20, end_seconds=26, interval_seconds=3),
            ],
        )

        plan = build_frame_plan(config, video)
        self.assertEqual([entry.seconds for entry in plan], [0, 4, 6, 8, 10, 20, 23, 26, 30])
        self.assertEqual([entry.is_dense for entry in plan], [False, True, True, True, False, True, True, True, False])

    def test_frame_sampler_estimate_counts_sheets(self) -> None:
        video = VideoInfo(Path("demo.mp4"), duration_seconds=1800, width=1200, height=2600, fps=30, file_size_bytes=1000)
        config = FrameSamplerConfig(
            video_path=Path("demo.mp4"),
            output_dir=Path("out"),
            start_seconds=0,
            end_seconds=1800,
            interval_seconds=10,
            sheet_cols=5,
            sheet_rows=6,
        )

        estimate = estimate_sampling(config, video)
        self.assertEqual(estimate.frame_count, 181)
        self.assertEqual(estimate.sheet_count, 7)

    def test_frame_sampler_loads_click_markers(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "mouse_events.jsonl"
            path.write_text(
                "\n".join(
                    [
                        json.dumps({"event_type": "move_sample", "t_video_ms": 100, "video_x": 1, "video_y": 2}),
                        json.dumps({"event_type": "click", "t_video_ms": 1200, "video_x": 30, "video_y": 40}),
                    ]
                ),
                encoding="utf-8",
            )

            markers = load_click_markers(path)

        self.assertEqual(len(markers), 1)
        self.assertEqual(markers[0].seconds, 1.2)
        self.assertEqual((markers[0].x, markers[0].y), (30, 40))

    def test_frame_sampler_click_marker_uses_half_second_window(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "mouse_events.jsonl"
            path.write_text(
                "\n".join(
                    [
                        json.dumps({"event_type": "click", "t_video_ms": 1000, "video_x": 30, "video_y": 40}),
                        json.dumps({"event_type": "click", "t_video_ms": 1700, "video_x": 50, "video_y": 60}),
                    ]
                ),
                encoding="utf-8",
            )
            markers = load_click_markers(path)

        self.assertIsNotNone(_nearest_click_marker(markers, 1.45, 0.5))
        self.assertIsNone(_nearest_click_marker(markers, 2.3, 0.5))

    def test_frame_sampler_maps_click_marker_after_crop(self) -> None:
        image = Image.new("RGB", (100, 100), "white")
        marker = load_click_markers(None)
        self.assertEqual(marker, [])

        thumb, position = _prepare_frame_image(
            image,
            CropRegion(x=20, y=10, width=40, height=40),
            200,
            ClickMarker(seconds=1.0, x=30, y=20),
        )

        self.assertEqual(thumb.size, (200, 200))
        self.assertEqual(position, (50, 50))

    def test_frame_sampler_labels_stay_outside_cropped_image(self) -> None:
        thumb = Image.new("RGB", (160, 24), "red")
        entry = FramePlanEntry(
            index=1,
            seconds=1.0,
            timestamp="00:00:01",
            is_dense=False,
            sheet_index=1,
            sheet_row=1,
            sheet_col=1,
        )
        config = FrameSamplerConfig(
            video_path=Path("video.mp4"),
            output_dir=Path("out"),
            sheet_cols=1,
            sheet_rows=1,
            show_timestamp=True,
            show_index=True,
        )

        sheet = _compose_sheet([thumb], [entry], config)

        self.assertEqual(sheet.getpixel((18, 64)), (255, 0, 0))
        self.assertNotEqual(sheet.getpixel((18, 82)), (255, 0, 0))

    def test_click_keyframes_cluster_dedupe_keeps_head_for_short_cluster_without_visual(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    [
                        json.dumps({"event_id": "a", "event_type": "click", "t_video_ms": 1000, "video_x": 10, "video_y": 10}),
                        json.dumps({"event_id": "b", "event_type": "click", "t_video_ms": 1200, "video_x": 18, "video_y": 14}),
                        json.dumps({"event_id": "c", "event_type": "click", "t_video_ms": 1400, "video_x": 19, "video_y": 14}),
                        json.dumps({"event_id": "d", "event_type": "click", "t_video_ms": 1800, "video_x": 18, "video_y": 14}),
                        json.dumps(
                            {
                                "event_id": "e",
                                "event_type": "double_click_candidate",
                                "t_video_ms": 1900,
                                "video_x": 18,
                                "video_y": 14,
                            }
                        ),
                    ]
                ),
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                time_dedupe_seconds=1.0,
                distance_dedupe_px=20,
            )

            events = load_click_keyframe_events(config)
            selected, skipped = select_click_keyframes(events, config)

        self.assertEqual([event.event_id for event in events], ["a", "b", "c", "d"])
        self.assertEqual([event.event_id for event in selected], ["a"])
        self.assertEqual(skipped, 3)

    def test_click_keyframes_cluster_dedupe_keeps_tail_for_large_cluster(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    json.dumps(
                        {
                            "event_id": f"evt_{index}",
                            "event_type": "click",
                            "t_video_ms": 1000 + index * 100,
                            "video_x": 10 + index,
                            "video_y": 10,
                        }
                    )
                    for index in range(5)
                )
                + "\n",
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                time_dedupe_seconds=1.0,
                distance_dedupe_px=20,
                cluster_tail_min_size=5,
            )

            events = load_click_keyframe_events(config)
            selection = select_click_keyframes_with_stats(events, config)

        self.assertEqual([event.event_id for event in selection.events], ["evt_0", "evt_4"])
        self.assertEqual(selection.skipped_count, 3)
        self.assertEqual(selection.stats["cluster_tail_kept"], 1)

    def test_click_keyframes_still_respects_max_frames(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    json.dumps({"event_id": f"evt_{index}", "event_type": "click", "t_video_ms": index * 100, "video_x": index})
                    for index in range(1, 5)
                )
                + "\n",
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                max_frames=2,
                time_dedupe_seconds=0,
                distance_dedupe_px=0,
            )

            selected, skipped = select_click_keyframes(load_click_keyframe_events(config), config)

        self.assertEqual([event.event_id for event in selected], ["evt_1", "evt_4"])
        self.assertEqual(skipped, 2)

    def test_click_keyframe_cap_spans_the_full_timeline(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    json.dumps({"event_id": f"evt_{index}", "event_type": "click", "t_video_ms": index * 1000, "video_x": index})
                    for index in range(1, 11)
                )
                + "\n",
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                max_frames=3,
                time_dedupe_seconds=0,
                distance_dedupe_px=0,
            )

            selection = select_click_keyframes_with_stats(load_click_keyframe_events(config), config)

        self.assertEqual([event.event_id for event in selection.events], ["evt_1", "evt_5", "evt_10"])
        self.assertEqual(selection.stats["cap_strategy"], "uniform_timeline")

    def test_click_keyframe_estimate_counts_selected_events_and_sheets(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    json.dumps(
                        {
                            "event_id": f"evt_{index}",
                            "event_type": "click",
                            "t_video_ms": 1000 + index * 100,
                            "video_x": 10 + index,
                            "video_y": 10,
                        }
                    )
                    for index in range(5)
                )
                + "\n",
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                sheet_cols=1,
                sheet_rows=1,
                time_dedupe_seconds=1.0,
                distance_dedupe_px=20,
                cluster_tail_min_size=5,
            )

            estimate = estimate_click_keyframe_sampling(config)

        self.assertEqual(estimate.events_total, 5)
        self.assertEqual(estimate.events_kept, 2)
        self.assertEqual(estimate.events_skipped, 3)
        self.assertEqual(estimate.sheet_count, 2)
        self.assertEqual(estimate.visual_signature_frames, 5)
        self.assertEqual(estimate.cached_frame_reuses, 2)
        self.assertEqual(estimate.estimated_frame_extractions, 5)

    def test_visual_signature_frames_are_cached_for_sheet_reuse(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            events_path = root / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    json.dumps({"event_id": f"evt_{index}", "event_type": "click", "t_video_ms": 1000 + index * 100, "video_x": 20, "video_y": 20})
                    for index in range(3)
                )
                + "\n",
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=root / "recording.mp4",
                events_path=events_path,
                output_dir=root,
                time_dedupe_seconds=1.0,
                distance_dedupe_px=20,
                visual_change_threshold=0.22,
            )
            events = load_click_keyframe_events(config)
            cache: dict[str, Path] = {}
            video = VideoInfo(config.video_path, duration_seconds=5, width=100, height=100, fps=30, file_size_bytes=100)
            with patch(
                "screen_mouse_recorder.frame_export.click_keyframes._extract_frame",
                side_effect=lambda *_args: Image.new("RGB", (100, 100), "white"),
            ):
                signatures = build_click_keyframe_visual_signatures(
                    events,
                    config,
                    video,
                    "ffmpeg",
                    frame_cache=cache,
                    frame_cache_dir=root / "cache",
                )

            self.assertEqual(set(signatures), {"evt_0", "evt_1", "evt_2"})
            self.assertEqual(set(cache), set(signatures))
            self.assertTrue(all(path.is_file() for path in cache.values()))

    def test_run_frame_export_dispatches_by_config_type(self) -> None:
        interval_config = FrameSamplerConfig(video_path=Path("video.mp4"), output_dir=Path("out"))
        click_config = ClickKeyframeConfig(
            video_path=Path("video.mp4"),
            events_path=Path("mouse_events.jsonl"),
            output_dir=Path("out"),
        )

        with patch("screen_mouse_recorder.frame_export.service.sample_video_to_sheets", return_value="interval") as interval_run:
            self.assertEqual(run_frame_export(interval_config, "ffmpeg"), "interval")
            interval_run.assert_called_once()

        with patch("screen_mouse_recorder.frame_export.service.generate_click_keyframe_sheets", return_value="click") as click_run:
            self.assertEqual(run_frame_export(click_config, "ffmpeg"), "click")
            click_run.assert_called_once()

    def test_click_keyframes_adds_silent_gap_compensation(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    [
                        json.dumps({"event_id": "a", "event_type": "click", "t_video_ms": 10000, "video_x": 10, "video_y": 10}),
                        json.dumps({"event_id": "b", "event_type": "click", "t_video_ms": 50000, "video_x": 200, "video_y": 200}),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                time_dedupe_seconds=1.5,
                distance_dedupe_px=80,
                silent_gap_seconds=10,
                silent_long_gap_seconds=25,
                silent_max_frames_per_gap=2,
            )
            video = VideoInfo(Path("recording.mp4"), duration_seconds=60, width=300, height=500, fps=30, file_size_bytes=1000)

            events = load_click_keyframe_events(config)
            selection = select_click_keyframes_with_stats(events, config)
            combined = _add_silent_gap_keyframes(selection.events, config, video, selection)

        self.assertEqual([event.event_type for event in combined], ["click", "silent_gap", "silent_gap", "click"])
        self.assertEqual([round(event.seconds, 3) for event in combined], [10.0, 23.333, 36.667, 50.0])
        self.assertEqual(selection.stats["silent_gap_frames_added"], 2)
        self.assertEqual(selection.stats["timeline_max_gap_before_seconds"], 40)
        self.assertLess(selection.stats["timeline_max_gap_after_seconds"], 14)

    def test_silent_gap_frames_respect_the_frame_cap(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    [
                        json.dumps({"event_id": "a", "event_type": "click", "t_video_ms": 10000, "video_x": 10, "video_y": 10}),
                        json.dumps({"event_id": "b", "event_type": "click", "t_video_ms": 50000, "video_x": 200, "video_y": 200}),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                max_frames=2,
                time_dedupe_seconds=0,
                distance_dedupe_px=0,
                silent_gap_seconds=10,
                silent_max_frames_per_gap=5,
            )
            video = VideoInfo(Path("recording.mp4"), duration_seconds=60, width=300, height=500, fps=30, file_size_bytes=1000)
            selection = select_click_keyframes_with_stats(load_click_keyframe_events(config), config)
            combined = _add_silent_gap_keyframes(selection.events, config, video, selection)

        self.assertEqual([event.event_id for event in combined], ["a", "b"])
        self.assertEqual(selection.stats["silent_gap_frames_added"], 0)

    def test_click_keyframes_can_include_double_click_and_paginate_plan(self) -> None:
        with TemporaryDirectory() as directory:
            events_path = Path(directory) / "mouse_events.jsonl"
            events_path.write_text(
                "\n".join(
                    [
                        json.dumps({"event_id": "a", "event_type": "click", "t_video_ms": 1000, "video_x": 10, "video_y": 10}),
                        json.dumps(
                            {
                                "event_id": "b",
                                "event_type": "double_click_candidate",
                                "video_timecode": "00:00:02.000",
                                "video_x": 80,
                                "video_y": 40,
                            }
                        ),
                    ]
                ),
                encoding="utf-8",
            )
            config = ClickKeyframeConfig(
                video_path=Path("recording.mp4"),
                events_path=events_path,
                output_dir=Path(directory),
                include_double_clicks=True,
                sheet_cols=1,
                sheet_rows=1,
            )
            video = VideoInfo(Path("recording.mp4"), duration_seconds=5, width=100, height=100, fps=30, file_size_bytes=100)
            plan = build_click_keyframe_plan(load_click_keyframe_events(config), config, video)

        self.assertEqual([entry.event_id for entry in plan], ["a", "b"])
        self.assertEqual([entry.sheet_index for entry in plan], [1, 2])
        self.assertEqual(plan[1].timestamp, "00:00:02")

    def test_video_recorder_draws_mouse_cursor(self) -> None:
        region = Region(screen_x=10, screen_y=20, width=320, height=240)
        command = FFmpegRecorder._build_command("ffmpeg", region, Path("recording.mp4"), 30)
        draw_mouse_index = command.index("-draw_mouse")

        self.assertEqual(command[draw_mouse_index + 1], "1")

    def test_calibration_rejects_obvious_misclick(self) -> None:
        region = Region(screen_x=100, screen_y=200, width=400, height=300)
        clicks = self._calibration_clicks(region, offset_x=120, offset_y=0, event_offset_x=120, event_offset_y=0)

        result = build_calibration_result(region, clicks, click_tolerance_px=80)

        self.assertFalse(result["completed"])
        self.assertIn("可能点错检查区域", "；".join(result["failure_reasons"]))

    def test_video_coordinates_stay_raw_with_coordinate_check_data(self) -> None:
        with TemporaryDirectory() as directory:
            region = Region(screen_x=100, screen_y=200, width=400, height=300)
            calibration = build_calibration_result(region, self._calibration_clicks(region, offset_x=4, offset_y=-2), 80)
            timing = TimingContext(session_id="test", logger_start_monotonic_ms=1000, video_zero_monotonic_ms=1000)
            logger = MouseActivityLogger(
                region,
                timing,
                AppConfig(),
                JsonlWriter(Path(directory) / "events.jsonl"),
                JsonlWriter(Path(directory) / "samples.jsonl"),
                calibration_data=calibration,
            )

            row = logger._make_event("click", 104, 208, source="test")

            self.assertEqual(row["region_x"], 4)
            self.assertEqual(row["region_y"], 8)
            self.assertEqual(row["video_x"], 4)
            self.assertEqual(row["video_y"], 8)
            self.assertFalse(row["calibration_applied"])
            self.assertEqual(row["calibration_method"], "raw_video_region")
            self.assertTrue(row["coordinate_check_completed"])

    def test_coordinate_check_mapping_is_diagnostic_only(self) -> None:
        with TemporaryDirectory() as directory:
            region = Region(screen_x=100, screen_y=200, width=400, height=300)
            calibration = build_calibration_result(
                region,
                self._scaled_calibration_clicks(region, scale_x=1.1, scale_y=0.9, shift_x=15, shift_y=-8),
                click_tolerance_px=80,
            )
            timing = TimingContext(session_id="test", logger_start_monotonic_ms=1000, video_zero_monotonic_ms=1000)
            logger = MouseActivityLogger(
                region,
                timing,
                AppConfig(),
                JsonlWriter(Path(directory) / "events.jsonl"),
                JsonlWriter(Path(directory) / "samples.jsonl"),
                calibration_data=calibration,
            )

            row = logger._make_event("click", 335, 327, source="test")

            self.assertTrue(calibration["completed"])
            self.assertEqual(calibration["mapping"]["method"], "visual_affine_least_squares")
            self.assertFalse(calibration["mapping"]["applied_to_recording_rows"])
            self.assertEqual(row["video_x"], 235)
            self.assertEqual(row["video_y"], 127)
            self.assertEqual(row["calibration_method"], "raw_video_region")

    def test_outside_filter_uses_raw_video_coordinates(self) -> None:
        with TemporaryDirectory() as directory:
            region = Region(screen_x=100, screen_y=100, width=100, height=100)
            calibration = build_calibration_result(region, self._calibration_clicks(region, offset_x=5, offset_y=0), 80)
            config = AppConfig(record_outside_region=False)
            timing = TimingContext(session_id="test", logger_start_monotonic_ms=1000, video_zero_monotonic_ms=1000)
            logger = MouseActivityLogger(
                region,
                timing,
                config,
                JsonlWriter(Path(directory) / "events.jsonl"),
                JsonlWriter(Path(directory) / "samples.jsonl"),
                calibration_data=calibration,
            )

            row = logger._make_event("click", 204, 150, source="test")
            logger._enqueue_event(row)

            self.assertFalse(row["inside_region"])
            self.assertFalse(row["inside_video_region"])
            self.assertEqual(logger.event_queue.qsize(), 0)

    def test_logger_ids_continue_across_segments(self) -> None:
        with TemporaryDirectory() as directory:
            region = Region(screen_x=0, screen_y=0, width=100, height=100)
            timing = TimingContext(session_id="test", logger_start_monotonic_ms=1000, video_zero_monotonic_ms=1000)
            first = MouseActivityLogger(
                region,
                timing,
                AppConfig(),
                JsonlWriter(Path(directory) / "events_1.jsonl"),
                JsonlWriter(Path(directory) / "samples_1.jsonl"),
            )
            first_event = first._make_event("click", 10, 10, source="test")
            first_sample = first._make_sample(10, 10, t_ms=1000)
            second = MouseActivityLogger(
                region,
                timing,
                AppConfig(),
                JsonlWriter(Path(directory) / "events_2.jsonl"),
                JsonlWriter(Path(directory) / "samples_2.jsonl"),
                event_counter_start=first.event_counter,
                sample_counter_start=first.sample_counter,
            )
            second_event = second._make_event("click", 20, 20, source="test")
            second_sample = second._make_sample(20, 20, t_ms=1100)

            self.assertEqual(first_event["event_id"], "evt_000001")
            self.assertEqual(first_sample["sample_id"], "smp_000001")
            self.assertEqual(second_event["event_id"], "evt_000002")
            self.assertEqual(second_sample["sample_id"], "smp_000002")

    def test_summary_and_xlsx(self) -> None:
        with TemporaryDirectory() as directory:
            storage = self._sample_storage(Path(directory))
            summary = generate_summary(storage)

            self.assertEqual(summary["clicks_total"], 1)
            self.assertEqual(summary["drag_count"], 1)
            self.assertTrue(storage.mouse_summary.exists())
            self.assertTrue(zipfile.is_zipfile(storage.mouse_summary_xlsx))
            self.assertTrue(zipfile.is_zipfile(storage.mouse_analysis_xlsx))
            with zipfile.ZipFile(storage.mouse_summary_xlsx) as archive:
                sheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
            self.assertIn("Mouse Events", sheet)
            self.assertIn("Mouse Samples", sheet)
            self.assertIn("video_x", sheet)
            with zipfile.ZipFile(storage.mouse_analysis_xlsx) as archive:
                analysis_sheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
            self.assertIn("录制概览", analysis_sheet)
            self.assertIn("操作明细", analysis_sheet)
            self.assertIn("视频X", analysis_sheet)

    def test_summary_prefers_video_region_flag(self) -> None:
        with TemporaryDirectory() as directory:
            storage = SessionStorage(Path(directory))
            storage.mouse_events.write_text(
                json.dumps(
                    {
                        "event_type": "click",
                        "t_video_ms": 10,
                        "inside_region": False,
                        "inside_video_region": True,
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            storage.mouse_samples.write_text("", encoding="utf-8")

            summary = generate_summary(storage)

            self.assertEqual(summary["clicks_inside_region"], 1)
            self.assertEqual(summary["clicks_outside_region"], 0)

    def test_cli_postprocess(self) -> None:
        with TemporaryDirectory() as directory:
            storage = self._sample_storage(Path(directory))

            result = subprocess.run(
                [
                    str(PYTHON),
                    str(ROOT / "run.py"),
                    "postprocess",
                    str(storage.session_dir),
                ],
                cwd=ROOT,
                text=True,
                capture_output=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertIn('"clicks_total": 1', result.stdout)
            self.assertTrue(storage.mouse_summary.exists())

    def test_behavior_report_from_session_folder(self) -> None:
        with TemporaryDirectory() as directory:
            storage = self._analysis_storage(Path(directory))

            result = generate_behavior_report(storage.session_dir)

            self.assertEqual(result.output_dir, (storage.session_dir / "auto_report").resolve())
            self.assertEqual(result.metrics["clicks_total"], 2)
            self.assertTrue(zipfile.is_zipfile(result.outputs["report"]))
            self.assertEqual(result.outputs["click_keyframes"].name, "keyframes_click_sheet.png")
            self.assertTrue(result.outputs["heatmap_circle"].exists())
            self.assertFalse((result.output_dir / "click_heatmap_true_ratio.png").exists())
            self.assertFalse((result.output_dir / "click_heatmap_square_matrix.png").exists())
            workbook = load_workbook(result.outputs["report"], read_only=True)
            heatmap_sheet = workbook["点击热力图"]
            self.assertEqual(heatmap_sheet.max_row, 81)
            self.assertEqual(heatmap_sheet.max_column, 41)
            self.assertEqual(heatmap_sheet.cell(row=1, column=41).value, "列40")
            workbook.close()

    def test_behavior_report_from_summary_xlsx(self) -> None:
        with TemporaryDirectory() as directory:
            storage = self._analysis_storage(Path(directory))
            generate_summary(storage)

            result = generate_behavior_report(storage.mouse_summary_xlsx)

            self.assertEqual(result.output_dir, (storage.session_dir / "auto_report").resolve())
            self.assertEqual(result.metrics["clicks_total"], 2)
            self.assertTrue(zipfile.is_zipfile(result.outputs["report"]))

    def test_config_example_json_matches_dataclass_fields(self) -> None:
        from dataclasses import fields

        example = json.loads((ROOT / "config.example.json").read_text(encoding="utf-8-sig"))
        expected = {f.name for f in fields(AppConfig)}
        actual = set(example)

        self.assertEqual(
            actual,
            expected,
            "config.example.json is out of sync with AppConfig. "
            f"Missing: {sorted(expected - actual)}; Unexpected: {sorted(actual - expected)}",
        )

    @staticmethod
    def _sample_storage(session_dir: Path) -> SessionStorage:
        storage = SessionStorage(session_dir)
        storage.mouse_events.write_text(
            "\n".join(
                [
                    json.dumps({"event_type": "left_down", "t_video_ms": 10, "inside_region": True}),
                    json.dumps({"event_type": "left_up", "t_video_ms": 80, "inside_region": True}),
                    json.dumps({"event_type": "click", "t_video_ms": 80, "inside_region": True}),
                    json.dumps({"event_type": "wheel", "t_video_ms": 120, "inside_region": False}),
                    json.dumps({"event_type": "drag_start", "t_video_ms": 200, "inside_region": True}),
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        storage.mouse_samples.write_text(
            "\n".join(
                [
                    json.dumps({"event_type": "move_sample", "t_video_ms": 0}),
                    json.dumps({"event_type": "move_sample", "t_video_ms": 1000}),
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        return storage

    @staticmethod
    def _analysis_storage(session_dir: Path) -> SessionStorage:
        storage = SessionStorage(session_dir)
        storage.session_meta.write_text(
            json.dumps({"recording_region": {"screen_x": 0, "screen_y": 0, "width": 100, "height": 200}}),
            encoding="utf-8",
        )
        storage.mouse_events.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "event_id": "evt_000001",
                            "event_type": "click",
                            "t_video_ms": 100,
                            "video_timecode": "00:00.100",
                            "video_x": 20,
                            "video_y": 40,
                            "inside_video_region": True,
                        }
                    ),
                    json.dumps(
                        {
                            "event_id": "evt_000002",
                            "event_type": "click",
                            "t_video_ms": 500,
                            "video_timecode": "00:00.500",
                            "video_x": 50,
                            "video_y": 160,
                            "inside_video_region": True,
                        }
                    ),
                    json.dumps(
                        {
                            "event_id": "evt_000003",
                            "event_type": "double_click_candidate",
                            "t_video_ms": 520,
                            "video_timecode": "00:00.520",
                            "video_x": 50,
                            "video_y": 160,
                            "inside_video_region": True,
                        }
                    ),
                    json.dumps(
                        {
                            "event_id": "evt_000004",
                            "event_type": "drag_start",
                            "t_video_ms": 1000,
                            "video_timecode": "00:01.000",
                            "video_x": 10,
                            "video_y": 20,
                            "inside_video_region": True,
                        }
                    ),
                    json.dumps(
                        {
                            "event_id": "evt_000005",
                            "event_type": "drag_end",
                            "t_video_ms": 1300,
                            "video_timecode": "00:01.300",
                            "video_x": 70,
                            "video_y": 80,
                            "inside_video_region": True,
                        }
                    ),
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        storage.mouse_samples.write_text(
            "\n".join(
                [
                    json.dumps({"sample_id": "smp_000001", "t_video_ms": 0, "video_x": 10, "video_y": 20}),
                    json.dumps({"sample_id": "smp_000002", "t_video_ms": 2000, "video_x": 20, "video_y": 30}),
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        return storage

    @staticmethod
    def _calibration_clicks(
        region: Region,
        offset_x: int,
        offset_y: int,
        event_offset_x: int = 0,
        event_offset_y: int = 0,
    ) -> list[dict[str, int | str]]:
        inset = 10
        targets = [
            ("top_left", region.screen_x + inset, region.screen_y + inset),
            ("top_right", region.screen_x + region.width - inset, region.screen_y + inset),
            ("bottom_left", region.screen_x + inset, region.screen_y + region.height - inset),
            ("bottom_right", region.screen_x + region.width - inset, region.screen_y + region.height - inset),
            ("center", region.screen_x + region.width // 2, region.screen_y + region.height // 2),
        ]
        return [
            {
                "target_id": target_id,
                "label": target_id,
                "expected_screen_x": x,
                "expected_screen_y": y,
                "actual_screen_x": x + offset_x,
                "actual_screen_y": y + offset_y,
                "tk_event_screen_x": x + event_offset_x,
                "tk_event_screen_y": y + event_offset_y,
            }
            for target_id, x, y in targets
        ]

    @staticmethod
    def _scaled_calibration_clicks(
        region: Region,
        scale_x: float,
        scale_y: float,
        shift_x: float,
        shift_y: float,
    ) -> list[dict[str, int | str]]:
        inset = 10
        targets = [
            ("top_left", inset, inset),
            ("top_right", region.width - inset, inset),
            ("bottom_left", inset, region.height - inset),
            ("bottom_right", region.width - inset, region.height - inset),
            ("center", region.width // 2, region.height // 2),
        ]
        rows = []
        for target_id, video_x, video_y in targets:
            expected_screen_x = region.screen_x + video_x
            expected_screen_y = region.screen_y + video_y
            actual_screen_x = int(round(region.screen_x + video_x * scale_x + shift_x))
            actual_screen_y = int(round(region.screen_y + video_y * scale_y + shift_y))
            rows.append(
                {
                    "target_id": target_id,
                    "label": target_id,
                    "expected_screen_x": expected_screen_x,
                    "expected_screen_y": expected_screen_y,
                    "actual_screen_x": actual_screen_x,
                    "actual_screen_y": actual_screen_y,
                    "tk_event_screen_x": expected_screen_x,
                    "tk_event_screen_y": expected_screen_y,
                }
            )
        return rows


if __name__ == "__main__":
    unittest.main()
