from __future__ import annotations

from dataclasses import dataclass
from importlib.metadata import PackageNotFoundError, version
import json
from pathlib import Path
import re
import time
from typing import Any, Callable, Protocol

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw

from ..media_utils import extract_frame, parse_timecode, resolve_ffmpeg


ProgressCallback = Callable[[int, int, str], None]


@dataclass(slots=True)
class OCRTextItem:
    text: str
    confidence: float
    box: tuple[int, int, int, int] | None = None


class OCREngine(Protocol):
    name: str
    version: str

    def recognize(self, image_path: Path) -> tuple[list[OCRTextItem], float]: ...


@dataclass(slots=True)
class OCREventExtractionConfig:
    index_json: Path
    selection_json: Path
    output_dir: Path
    video_path: Path | None = None
    ffmpeg_path: str | None = None
    write_review_images: bool = True


@dataclass(slots=True)
class OCREventExtractionResult:
    output_dir: Path
    json_path: Path
    xlsx_path: Path
    review_dir: Path
    event_count: int
    needs_review_count: int
    elapsed_seconds: float


class RapidOCREngine:
    name = "rapidocr-onnxruntime"

    def __init__(self) -> None:
        try:
            from rapidocr_onnxruntime import RapidOCR
        except ImportError as exc:
            raise RuntimeError(
                'RapidOCR is not installed. Install the optional OCR dependencies with: '
                'python -m pip install -e ".[ocr]"'
            ) from exc
        try:
            self.version = version("rapidocr-onnxruntime")
        except PackageNotFoundError:
            self.version = "unknown"
        self._engine = RapidOCR()

    def recognize(self, image_path: Path) -> tuple[list[OCRTextItem], float]:
        started = time.perf_counter()
        raw, _engine_elapsed = self._engine(str(image_path))
        elapsed = time.perf_counter() - started
        items: list[OCRTextItem] = []
        for item in raw or []:
            if len(item) != 3:
                continue
            points, text, confidence = item
            xs = [int(round(point[0])) for point in points]
            ys = [int(round(point[1])) for point in points]
            items.append(
                OCRTextItem(
                    text=str(text),
                    confidence=float(confidence),
                    box=(min(xs), min(ys), max(xs), max(ys)),
                )
            )
        return items, elapsed


def extract_selected_ocr_events(
    config: OCREventExtractionConfig,
    *,
    engine: OCREngine | None = None,
    progress: ProgressCallback | None = None,
) -> OCREventExtractionResult:
    started = time.perf_counter()
    index_path = config.index_json.resolve()
    selection_path = config.selection_json.resolve()
    index_data = _read_json_object(index_path)
    selection_data = _read_json_object(selection_path)
    frames = index_data.get("frames")
    selections = selection_data.get("selections")
    if not isinstance(frames, list):
        raise ValueError(f"Index JSON has no frames list: {index_path}")
    if not isinstance(selections, list) or not selections:
        raise ValueError(f"Selection JSON has no selections: {selection_path}")

    ocr_engine = engine or RapidOCREngine()
    output_dir = config.output_dir.resolve()
    review_dir = output_dir / "ocr_review"
    output_dir.mkdir(parents=True, exist_ok=True)
    review_dir.mkdir(parents=True, exist_ok=True)

    video_path = _resolve_optional_path(
        config.video_path or selection_data.get("source_video"),
        selection_path.parent,
    )
    resolved = [_resolve_selection(selection, frames, index_path.parent) for selection in selections]
    needs_video_extract = any(
        not (selection.get("source_frame") or frame.get("source_frame")) for selection, frame in resolved
    )
    ffmpeg = resolve_ffmpeg(config.ffmpeg_path) if video_path is not None and needs_video_extract else None
    events: list[dict[str, Any]] = []
    total = len(resolved)
    for position, (selection, frame) in enumerate(resolved, start=1):
        event_id = f"ocr_evt_{position:06d}"
        if progress:
            progress(position - 1, total, f"OCR {frame['timestamp']} #{frame['index']:03d}")
        source_path = _prepare_source_image(
            event_id,
            selection,
            frame,
            index_path.parent,
            selection_path.parent,
            video_path,
            ffmpeg,
            review_dir,
        )
        items, ocr_elapsed = ocr_engine.recognize(source_path)
        review_path = review_dir / f"{event_id}_boxes.jpg"
        if config.write_review_images:
            _write_annotated_image(source_path, review_path, items)
        event = _build_event(
            event_id,
            selection,
            frame,
            index_path.parent,
            source_path,
            video_path,
            review_path if config.write_review_images else None,
            items,
            ocr_elapsed,
        )
        events.append(event)

    payload = {
        "schema_version": "1.0",
        "workflow": "manual_selected_frame_ocr",
        "source": {
            "video": str(video_path or ""),
            "index_json": str(index_path),
            "selection_json": str(selection_path),
            "ocr_engine": ocr_engine.name,
            "ocr_engine_version": ocr_engine.version,
        },
        "events": events,
    }
    json_path = output_dir / "event_ocr_results.json"
    xlsx_path = output_dir / "event_ocr_results.xlsx"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")
    _write_xlsx(xlsx_path, events)
    if progress:
        progress(total, total, "OCR 完成")
    return OCREventExtractionResult(
        output_dir=output_dir,
        json_path=json_path,
        xlsx_path=xlsx_path,
        review_dir=review_dir,
        event_count=len(events),
        needs_review_count=sum(event["review_status"] == "needs_review" for event in events),
        elapsed_seconds=round(time.perf_counter() - started, 4),
    )


def _read_json_object(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(path)
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(data, dict):
        raise ValueError(f"Expected a JSON object: {path}")
    return data


def _resolve_selection(
    selection: Any,
    frames: list[Any],
    index_dir: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    if not isinstance(selection, dict):
        raise ValueError("Each OCR selection must be an object.")
    tile_index = _safe_int(selection.get("tile_index"))
    if tile_index <= 0:
        raise ValueError(f"Invalid tile_index: {selection.get('tile_index')}")
    requested_sheet = str(selection.get("sheet") or "").strip()
    matches = []
    for frame in frames:
        if not isinstance(frame, dict) or _safe_int(frame.get("index")) != tile_index:
            continue
        if requested_sheet and Path(str(frame.get("sheet") or "")).name != Path(requested_sheet).name:
            continue
        matches.append(frame)
    if not matches:
        raise ValueError(f"Selected tile was not found in index: #{tile_index:03d} {requested_sheet}".strip())
    if len(matches) > 1 and not requested_sheet:
        names = ", ".join(str(frame.get("sheet") or "") for frame in matches)
        raise ValueError(f"Tile #{tile_index:03d} is ambiguous; specify sheet. Matches: {names}")
    frame = dict(matches[0])
    frame["index"] = tile_index
    frame["seconds"] = float(frame.get("seconds") or 0.0)
    frame["timestamp"] = str(frame.get("timestamp") or _format_seconds(frame["seconds"]))
    same_sheet = [
        candidate
        for candidate in frames
        if isinstance(candidate, dict) and str(candidate.get("sheet") or "") == str(frame.get("sheet") or "")
    ]
    frame["sheet_cols"] = max(5, max((_safe_int(candidate.get("sheet_col")) for candidate in same_sheet), default=0))
    frame["sheet_rows"] = max(6, max((_safe_int(candidate.get("sheet_row")) for candidate in same_sheet), default=0))
    return dict(selection), frame


def _prepare_source_image(
    event_id: str,
    selection: dict[str, Any],
    frame: dict[str, Any],
    index_dir: Path,
    selection_dir: Path,
    video_path: Path | None,
    ffmpeg: str | None,
    review_dir: Path,
) -> Path:
    source_value = selection.get("source_frame") or frame.get("source_frame")
    source_frame = _resolve_optional_path(source_value, selection_dir)
    target = review_dir / f"{event_id}_source.jpg"
    if source_frame is not None:
        if not source_frame.exists():
            raise FileNotFoundError(source_frame)
        _save_as_jpeg(source_frame, target)
        return target
    if video_path is not None:
        if ffmpeg is None:
            raise RuntimeError("FFmpeg is required to extract OCR source frames from video.")
        image = extract_frame(ffmpeg, video_path, float(frame["seconds"]))
        try:
            image.save(target, "JPEG", quality=95)
        finally:
            image.close()
        return target
    sheet_value = str(frame.get("sheet") or "").strip()
    sheet_path = _resolve_sheet_path(sheet_value, index_dir) if sheet_value else None
    if sheet_path is not None and sheet_path.is_file():
        _crop_contact_sheet_tile(sheet_path, frame, target)
        return target
    raise ValueError(
        f"No OCR source available for tile #{frame['index']:03d}. Provide source_frame, video_path, or contact sheet."
    )


def _crop_contact_sheet_tile(sheet_path: Path, frame: dict[str, Any], target: Path) -> None:
    with Image.open(sheet_path) as sheet:
        sheet = sheet.convert("RGB")
        row = max(1, _safe_int(frame.get("sheet_row")))
        col = max(1, _safe_int(frame.get("sheet_col")))
        cols = max(col, _safe_int(frame.get("sheet_cols")) or 5)
        rows = max(row, _safe_int(frame.get("sheet_rows")) or 6)
        gap = 8
        header_h = 46
        thumb_w = max(1, (sheet.width - (cols + 1) * gap) // cols)
        label_h = max(22, min(34, thumb_w // 11))
        thumb_h = max(1, (sheet.height - header_h - (rows + 1) * gap) // rows - label_h)
        left = gap + (col - 1) * (thumb_w + gap)
        top = header_h + gap + (row - 1) * (thumb_h + label_h + gap)
        crop = sheet.crop((left, top, left + thumb_w, top + thumb_h))
        crop.save(target, "JPEG", quality=95)


def _build_event(
    event_id: str,
    selection: dict[str, Any],
    frame: dict[str, Any],
    index_dir: Path,
    source_path: Path,
    video_path: Path | None,
    review_path: Path | None,
    items: list[OCRTextItem],
    ocr_elapsed: float,
) -> dict[str, Any]:
    raw_text = " ".join(item.text.strip() for item in items if item.text.strip())
    normalized = _normalize_text(raw_text)
    event_type = _classify_event(normalized)
    event_name = _extract_event_name(event_type, items, normalized, str(selection.get("note") or ""))
    ocr_time = _extract_ocr_time(raw_text)
    time_check = _check_time(str(frame["timestamp"]), ocr_time)
    confidence = _event_confidence(items, event_type, event_name)
    review_status = "pending"
    if event_type == "unknown" or not event_name or time_check == "ocr_mismatch":
        review_status = "needs_review"
    return {
        "event_id": event_id,
        "event_type": event_type,
        "event_name": event_name or "待人工确认",
        "timestamp": str(frame["timestamp"]),
        "seconds": round(float(frame["seconds"]), 3),
        "frame_index": int(frame["index"]),
        "source_frame": str(source_path),
        "source_video": str(video_path or ""),
        "contact_sheet": (
            str(_resolve_sheet_path(frame.get("sheet"), index_dir)) if str(frame.get("sheet") or "").strip() else ""
        ),
        "contact_sheet_tile": int(frame["index"]),
        "sheet_row": _safe_int(frame.get("sheet_row")),
        "sheet_col": _safe_int(frame.get("sheet_col")),
        "ocr_text": raw_text,
        "ocr_time_text": ocr_time,
        "time_source": "index_json",
        "time_check": time_check,
        "confidence": confidence,
        "review_status": review_status,
        "review_image": str(review_path or ""),
        "ocr_elapsed_seconds": round(float(ocr_elapsed), 4),
        "notes": str(selection.get("note") or ""),
        "ocr_items": [
            {
                "text": item.text,
                "confidence": round(item.confidence, 4),
                "box": list(item.box) if item.box is not None else None,
            }
            for item in items
        ],
    }


def _classify_event(normalized: str) -> str:
    if "新功能开启" in normalized:
        return "new_feature_unlocked"
    if "新技能解锁" in normalized or "技能解锁" in normalized:
        return "new_skill_unlocked"
    if "战力" in normalized and re.search(r"\d+(?:\.\d+)?万?", normalized):
        return "combat_power_snapshot"
    if re.search(r"\d+级", normalized) or "转生" in normalized:
        return "level_snapshot"
    if any(keyword in normalized for keyword in ("获得", "获取", "任务奖励", "领取")):
        return "reward_popup"
    if any(keyword in normalized for keyword in ("任务完成", "任务进度", "前往任务")):
        return "task_progress"
    return "unknown"


def _extract_event_name(event_type: str, items: list[OCRTextItem], normalized: str, note: str) -> str:
    if event_type == "new_feature_unlocked":
        return _near_marker_title(items, "新功能开启", direction="above") or _name_from_note(note)
    if event_type == "new_skill_unlocked":
        return _near_marker_title(items, "新技能解锁", direction="below") or _name_from_note(note)
    if event_type == "combat_power_snapshot":
        match = re.search(r"战力[:：]?([0-9]+(?:\.[0-9]+)?万?)", normalized)
        return match.group(1) if match else ""
    if event_type == "level_snapshot":
        match = re.search(r"([0-9]+级(?:未转生|已转生)?)", normalized)
        return match.group(1) if match else ""
    if event_type == "reward_popup":
        reward_items = [item.text.strip() for item in items if any(key in item.text for key in ("获得", "获取", "奖励"))]
        return "; ".join(reward_items[:3]) or _name_from_note(note)
    return _name_from_note(note)


def _near_marker_title(items: list[OCRTextItem], marker_text: str, *, direction: str) -> str:
    marker = next((item for item in items if marker_text in _normalize_text(item.text) and item.box), None)
    if marker is None or marker.box is None:
        return ""
    marker_left, marker_top, marker_right, marker_bottom = marker.box
    marker_center = (marker_left + marker_right) / 2
    candidates: list[tuple[float, OCRTextItem]] = []
    for item in items:
        if item is marker or item.box is None or not _is_title_candidate(item.text):
            continue
        left, top, right, bottom = item.box
        center = (left + right) / 2
        if direction == "above":
            gap = marker_top - bottom
        else:
            gap = top - marker_bottom
        if gap < -8 or gap > 220:
            continue
        horizontal_penalty = abs(center - marker_center) * 0.25
        candidates.append((gap + horizontal_penalty, item))
    if not candidates:
        return ""
    candidates.sort(key=lambda pair: pair[0])
    return candidates[0][1].text.strip()


def _is_title_candidate(text: str) -> bool:
    cleaned = text.strip()
    normalized = _normalize_text(cleaned)
    if not cleaned or len(cleaned) > 16 or re.fullmatch(r"[#\d:./+\-]+", cleaned):
        return False
    excluded = (
        "新功能开启",
        "新技能解锁",
        "任务奖励",
        "技能效果",
        "青云诀之伏魔",
        "点击空白",
        "世界",
    )
    return not any(value in normalized for value in excluded)


def _event_confidence(items: list[OCRTextItem], event_type: str, event_name: str) -> float:
    if not items:
        return 0.0
    relevant = []
    markers = {
        "new_feature_unlocked": "新功能开启",
        "new_skill_unlocked": "新技能解锁",
        "combat_power_snapshot": "战力",
    }
    marker = markers.get(event_type, "")
    for item in items:
        normalized = _normalize_text(item.text)
        if marker and marker in normalized or event_name and event_name in item.text:
            relevant.append(item.confidence)
    values = relevant or [item.confidence for item in items]
    return round(sum(values) / len(values), 4)


def _extract_ocr_time(text: str) -> str:
    match = re.search(r"(?<!\d)(\d{1,2}:\d{2}:\d{2})(?!\d)", text)
    return match.group(1) if match else ""


def _check_time(index_time: str, ocr_time: str) -> str:
    if not ocr_time:
        return "ocr_missing"
    try:
        expected = parse_timecode(index_time)
        actual = parse_timecode(ocr_time)
    except ValueError:
        return "ocr_mismatch"
    if expected is None or actual is None:
        return "ocr_mismatch"
    return "matched" if abs(expected - actual) <= 1.0 else "ocr_mismatch"


def _normalize_text(text: str) -> str:
    return (
        text.replace(" ", "")
        .replace("\n", "")
        .replace("成力", "战力")
        .replace("诚力", "战力")
        .replace("城力", "战力")
        .replace("戰力", "战力")
    )


def _name_from_note(note: str) -> str:
    text = note.strip()
    if not text:
        return ""
    for separator in ("：", ":"):
        if separator in text:
            return text.split(separator, 1)[1].strip()
    return text


def _write_annotated_image(source_path: Path, target: Path, items: list[OCRTextItem]) -> None:
    with Image.open(source_path) as source:
        image = source.convert("RGB")
    draw = ImageDraw.Draw(image)
    for item in items:
        if item.box is not None:
            draw.rectangle(item.box, outline=(255, 40, 40), width=2)
    image.save(target, "JPEG", quality=92)


def _save_as_jpeg(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(source) as image:
        image.convert("RGB").save(target, "JPEG", quality=95)


def _write_xlsx(path: Path, events: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OCR事件"
    headers = [
        "event_id",
        "event_type",
        "event_name",
        "timestamp",
        "seconds",
        "frame_index",
        "sheet_row",
        "sheet_col",
        "confidence",
        "review_status",
        "time_check",
        "ocr_text",
        "source_frame",
        "contact_sheet",
        "review_image",
        "notes",
    ]
    sheet.append(headers)
    for event in events:
        sheet.append([event.get(header, "") for header in headers])
    fill = PatternFill("solid", fgColor="E9EEF5")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    widths = {
        "A": 20,
        "B": 24,
        "C": 24,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 10,
        "H": 10,
        "I": 12,
        "J": 16,
        "K": 14,
        "L": 72,
        "M": 70,
        "N": 70,
        "O": 70,
        "P": 36,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def _resolve_optional_path(value: Any, base_dir: Path) -> Path | None:
    if value is None or str(value).strip() == "":
        return None
    path = Path(str(value))
    return path.resolve() if path.is_absolute() else (base_dir / path).resolve()


def _resolve_sheet_path(value: Any, base_dir: Path) -> Path:
    if value is None or str(value).strip() == "":
        return base_dir
    path = Path(str(value))
    return path.resolve() if path.is_absolute() else (base_dir / path).resolve()


def _safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _format_seconds(seconds: float) -> str:
    total = max(0, int(round(seconds)))
    hours, remainder = divmod(total, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"
