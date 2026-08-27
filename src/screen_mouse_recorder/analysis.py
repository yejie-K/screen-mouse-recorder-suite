from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
import json
import math
from pathlib import Path
from statistics import mean, median
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont

from .analysis_handoff import ANALYSIS_HANDOFF_RELATIVE_PATH, write_analysis_handoff
from .frame_sampler import build_click_summary_config, generate_click_keyframe_sheets
from .naming import default_report_output_dir
from .postprocess import iter_jsonl, row_inside_video_region
from .storage import SessionStorage


HEATMAP_SHORT_SIDE_CELLS = 40
COARSE_GRID_SIZE = 3


@dataclass(slots=True)
class BehaviorAnalysisResult:
    source_path: Path
    output_dir: Path
    metrics: dict[str, Any]
    outputs: dict[str, Path]
    warnings: list[str]


def default_analysis_output_dir(source_path: Path) -> Path:
    return default_report_output_dir(source_path)


def generate_behavior_report(
    source_path: Path,
    output_dir: Path | None = None,
    ffmpeg_path: str | None = None,
) -> BehaviorAnalysisResult:
    source_path = source_path.resolve()
    output_dir = (output_dir or default_analysis_output_dir(source_path)).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    events, samples, meta, warnings = _load_source_rows(source_path)
    if not events and not samples:
        raise ValueError("没有找到可分析的鼠标事件或采样数据。")

    region_width, region_height = _recording_region_size(events, samples, meta, warnings)
    analysis = _build_analysis(events, samples, region_width, region_height)

    outputs = {
        "metrics": output_dir / "metrics.json",
        "report": output_dir / "report_summary.xlsx",
        "timeline": output_dir / "chart_activity_timeline.png",
        "heatmap_circle": output_dir / "chart_click_heatmap.png",
        "scatter": output_dir / "chart_click_scatter.png",
        "drag_durations": output_dir / "chart_drag_duration.png",
        "click_keyframes": output_dir / "keyframes_click_sheet.png",
    }

    _cleanup_legacy_report_outputs(output_dir)
    outputs["metrics"].write_text(
        json.dumps(analysis["metrics"], ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    _draw_activity_timeline(outputs["timeline"], analysis)
    _draw_click_circle_heatmap(outputs["heatmap_circle"], analysis)
    _draw_click_scatter(outputs["scatter"], analysis)
    _draw_drag_durations(outputs["drag_durations"], analysis)
    _write_behavior_workbook(outputs["report"], analysis)
    _maybe_generate_click_keyframes(source_path, output_dir, warnings, ffmpeg_path)

    return BehaviorAnalysisResult(
        source_path=source_path,
        output_dir=output_dir,
        metrics=analysis["metrics"],
        outputs=outputs,
        warnings=warnings,
    )


def _cleanup_legacy_report_outputs(output_dir: Path) -> None:
    for filename in (
        "summary_metrics.json",
        "mouse_behavior_report.xlsx",
        "activity_timeline.png",
        "click_heatmap_circle.png",
        "click_heatmap_true_ratio.png",
        "click_heatmap_square_matrix.png",
        "click_scatter.png",
        "drag_durations.png",
        "click_keyframes.png",
        "click_keyframes_index.json",
    ):
        try:
            (output_dir / filename).unlink(missing_ok=True)
        except OSError:
            pass
    for path in output_dir.glob("click_keyframes_*.png"):
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def _maybe_generate_click_keyframes(
    source_path: Path,
    output_dir: Path,
    warnings: list[str],
    ffmpeg_path: str | None,
) -> None:
    session_dir = source_path if source_path.is_dir() else source_path.parent
    storage = SessionStorage(session_dir)
    if not storage.recording_mp4.exists() or not storage.mouse_events.exists():
        return
    handoff_path = session_dir / ANALYSIS_HANDOFF_RELATIVE_PATH
    handoff_path.unlink(missing_ok=True)
    try:
        result = generate_click_keyframe_sheets(
            build_click_summary_config(
                storage.recording_mp4,
                storage.mouse_events,
                output_dir,
            ),
            ffmpeg_path=ffmpeg_path,
        )
    except Exception as exc:
        warnings.append(f"点击关键帧图生成失败：{exc}")
        return
    if not result.sheet_paths:
        warnings.append("点击关键帧图未生成：无点击事件。")
        return
    try:
        write_analysis_handoff(
            session_dir,
            frame_index_path=result.index_json,
            contact_sheet_paths=result.sheet_paths,
        )
    except Exception as exc:
        warnings.append(f"分析交接清单生成失败：{exc}")
    if result.warnings:
        warnings.extend(f"点击关键帧图：{warning}" for warning in result.warnings)


def _load_source_rows(source_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any], list[str]]:
    warnings: list[str] = []
    if source_path.is_dir():
        storage = SessionStorage(source_path)
        meta = _load_json(storage.session_meta)
        events = iter_jsonl(storage.mouse_events)
        samples = iter_jsonl(storage.mouse_samples)
        if not events and storage.mouse_summary_xlsx.exists():
            xlsx_events, xlsx_samples = _read_mouse_summary_xlsx(storage.mouse_summary_xlsx)
            events = xlsx_events
            samples = xlsx_samples
        if not events and not samples:
            raise ValueError(f"目录中没有 mouse_events.jsonl、mouse_samples.jsonl 或 mouse_summary.xlsx：{source_path}")
        return events, samples, meta, warnings

    if source_path.suffix.lower() != ".xlsx":
        raise ValueError("请选择 session 文件夹或 .xlsx 文件。")

    events, samples = _read_mouse_summary_xlsx(source_path)
    meta = _load_json(source_path.parent / "session_meta.json")
    if not events and not samples:
        raise ValueError(f"xlsx 中没有识别到 Mouse Events / Mouse Samples 数据：{source_path}")
    if not meta:
        warnings.append("未找到同目录 session_meta.json，热力图尺寸将从坐标范围推断。")
    return events, samples, meta, warnings


def _load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return {}


def _read_mouse_summary_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        events: list[dict[str, Any]] = []
        samples: list[dict[str, Any]] = []
        for worksheet in wb.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            events.extend(_extract_section_rows(rows, "Mouse Events"))
            samples.extend(_extract_section_rows(rows, "Mouse Samples"))
            if not events:
                events.extend(_extract_chinese_operation_rows(rows))
        return events, samples
    finally:
        wb.close()


def _extract_section_rows(rows: list[list[Any]], section_name: str) -> list[dict[str, Any]]:
    for index, row in enumerate(rows):
        if row and row[0] == section_name:
            if index + 1 >= len(rows):
                return []
            columns = [str(value) if value is not None else "" for value in rows[index + 1]]
            extracted: list[dict[str, Any]] = []
            for data_row in rows[index + 2 :]:
                first = data_row[0] if data_row else None
                if first in {"Mouse Events", "Mouse Samples"}:
                    break
                if not any(value is not None for value in data_row):
                    break
                item = {column: value for column, value in zip(columns, data_row, strict=False) if column}
                extracted.append(item)
            return extracted
    return []


def _extract_chinese_operation_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    for index, row in enumerate(rows):
        if row and row[0] == "操作明细":
            if index + 1 >= len(rows):
                return []
            columns = [str(value) if value is not None else "" for value in rows[index + 1]]
            extracted: list[dict[str, Any]] = []
            for data_row in rows[index + 2 :]:
                if not any(value is not None for value in data_row):
                    break
                item = {column: value for column, value in zip(columns, data_row, strict=False) if column}
                event_type = item.get("原始事件") or item.get("操作类型")
                extracted.append(
                    {
                        "event_id": item.get("事件ID"),
                        "event_type": _normalize_chinese_event(event_type),
                        "video_timecode": item.get("视频时间"),
                        "video_x": item.get("视频X"),
                        "video_y": item.get("视频Y"),
                        "duration_ms": item.get("持续毫秒"),
                        "inside_video_region": item.get("是否在录制区域内") != "否",
                    }
                )
            return extracted
    return []


def _normalize_chinese_event(value: Any) -> str:
    labels = {
        "点击": "click",
        "双击候选": "double_click_candidate",
        "滚轮": "wheel",
        "拖拽开始": "drag_start",
        "拖拽结束": "drag_end",
    }
    return labels.get(str(value), str(value))


def _recording_region_size(
    events: list[dict[str, Any]],
    samples: list[dict[str, Any]],
    meta: dict[str, Any],
    warnings: list[str],
) -> tuple[int, int]:
    region = meta.get("recording_region") if isinstance(meta.get("recording_region"), dict) else {}
    width = _safe_int(region.get("width"), 0)
    height = _safe_int(region.get("height"), 0)
    if width > 0 and height > 0:
        return width, height

    xs: list[float] = []
    ys: list[float] = []
    for row in events + samples:
        x = _safe_float(row.get("video_x"))
        y = _safe_float(row.get("video_y"))
        if x is not None:
            xs.append(x)
        if y is not None:
            ys.append(y)
    width = max(1, int(math.ceil(max(xs, default=0) + 1)))
    height = max(1, int(math.ceil(max(ys, default=0) + 1)))
    warnings.append(f"录制区域尺寸由坐标范围推断为 {width}x{height}。")
    return width, height


def _build_analysis(
    events: list[dict[str, Any]],
    samples: list[dict[str, Any]],
    region_width: int,
    region_height: int,
) -> dict[str, Any]:
    event_counts = Counter(str(row.get("event_type", "unknown")) for row in events)
    clicks: list[tuple[float, float, float]] = []
    minute_counts: Counter[int] = Counter()
    minute_clicks: Counter[int] = Counter()
    minute_drags: Counter[int] = Counter()
    drag_pairs: list[dict[str, float]] = []
    drag_start: dict[str, Any] | None = None

    for row in events:
        t_ms = _safe_float(row.get("t_video_ms")) or _timecode_to_ms(row.get("video_timecode")) or 0.0
        minute = int(t_ms // 60_000)
        event_type = str(row.get("event_type", "unknown"))
        minute_counts[minute] += 1
        if event_type == "click":
            x = _safe_float(row.get("video_x"))
            y = _safe_float(row.get("video_y"))
            if x is not None and y is not None and row_inside_video_region(row):
                clicks.append((x, y, t_ms))
            minute_clicks[minute] += 1
        elif event_type == "drag_start":
            drag_start = row
        elif event_type == "drag_end":
            if drag_start is not None:
                start_ms = _safe_float(drag_start.get("t_video_ms")) or _timecode_to_ms(drag_start.get("video_timecode")) or t_ms
                start_x = _safe_float(drag_start.get("video_x")) or 0.0
                start_y = _safe_float(drag_start.get("video_y")) or 0.0
                end_x = _safe_float(row.get("video_x")) or start_x
                end_y = _safe_float(row.get("video_y")) or start_y
                distance = ((end_x - start_x) ** 2 + (end_y - start_y) ** 2) ** 0.5
                drag_pairs.append(
                    {
                        "start_ms": start_ms,
                        "end_ms": t_ms,
                        "duration_ms": max(0.0, t_ms - start_ms),
                        "distance_px": distance,
                    }
                )
            minute_drags[minute] += 1
            drag_start = None

    duration_ms = _duration_ms(events, samples)
    duration_minutes = duration_ms / 60_000 if duration_ms else 0.0
    heatmap_cols, heatmap_rows = _heatmap_grid_dimensions(region_width, region_height)
    click_grid_3 = [[0 for _ in range(COARSE_GRID_SIZE)] for _ in range(COARSE_GRID_SIZE)]
    heat_grid = [[0 for _ in range(heatmap_cols)] for _ in range(heatmap_rows)]
    for x, y, _ in clicks:
        gx3 = min(COARSE_GRID_SIZE - 1, max(0, int(x / max(1, region_width / COARSE_GRID_SIZE))))
        gy3 = min(COARSE_GRID_SIZE - 1, max(0, int(y / max(1, region_height / COARSE_GRID_SIZE))))
        click_grid_3[gy3][gx3] += 1
        gx = min(heatmap_cols - 1, max(0, int(x / region_width * heatmap_cols)))
        gy = min(heatmap_rows - 1, max(0, int(y / region_height * heatmap_rows)))
        heat_grid[gy][gx] += 1

    drag_durations = [item["duration_ms"] for item in drag_pairs]
    drag_distances = [item["distance_px"] for item in drag_pairs]
    bottom_clicks = sum(click_grid_3[2])
    bottom_center_clicks = click_grid_3[2][1]
    right_clicks = sum(row[2] for row in click_grid_3)
    sample_hz = len(samples) / (duration_ms / 1000) if duration_ms else 0
    clicks_total = event_counts.get("click", 0)
    metrics = {
        "duration_minutes": round(duration_minutes, 3),
        "events_total": len(events),
        "samples_total": len(samples),
        "clicks_total": clicks_total,
        "clicks_with_position": len(clicks),
        "clicks_per_minute": round(clicks_total / duration_minutes, 3) if duration_minutes else 0,
        "double_click_candidates": event_counts.get("double_click_candidate", 0),
        "double_click_candidate_ratio": round(event_counts.get("double_click_candidate", 0) / clicks_total, 3)
        if clicks_total
        else 0,
        "wheel_events": event_counts.get("wheel", 0),
        "drag_count": len(drag_pairs) or event_counts.get("drag_start", 0),
        "avg_drag_duration_ms": round(mean(drag_durations), 1) if drag_durations else 0,
        "median_drag_duration_ms": round(median(drag_durations), 1) if drag_durations else 0,
        "max_drag_duration_ms": round(max(drag_durations), 1) if drag_durations else 0,
        "avg_drag_distance_px": round(mean(drag_distances), 1) if drag_distances else 0,
        "bottom_third_click_ratio": round(bottom_clicks / len(clicks), 3) if clicks else 0,
        "bottom_center_click_ratio": round(bottom_center_clicks / len(clicks), 3) if clicks else 0,
        "right_third_click_ratio": round(right_clicks / len(clicks), 3) if clicks else 0,
        "actual_sample_hz": round(sample_hz, 2),
        "region_width": region_width,
        "region_height": region_height,
        "heatmap_columns": heatmap_cols,
        "heatmap_rows": heatmap_rows,
        "event_counts": dict(sorted(event_counts.items())),
    }
    return {
        "metrics": metrics,
        "clicks": clicks,
        "drag_pairs": drag_pairs,
        "minutes": list(range(0, max(1, int(math.ceil(duration_minutes))))),
        "minute_counts": minute_counts,
        "minute_clicks": minute_clicks,
        "minute_drags": minute_drags,
        "click_grid_3": click_grid_3,
        "heat_grid": heat_grid,
        "region_width": region_width,
        "region_height": region_height,
    }


def _heatmap_grid_dimensions(region_width: int, region_height: int) -> tuple[int, int]:
    if region_width <= 0 or region_height <= 0:
        return HEATMAP_SHORT_SIDE_CELLS, HEATMAP_SHORT_SIDE_CELLS
    if region_width >= region_height:
        rows = HEATMAP_SHORT_SIDE_CELLS
        cols = max(1, round(region_width / region_height * rows))
    else:
        cols = HEATMAP_SHORT_SIDE_CELLS
        rows = max(1, round(region_height / region_width * cols))
    return cols, rows


def _duration_ms(events: list[dict[str, Any]], samples: list[dict[str, Any]]) -> float:
    times = []
    for row in events + samples:
        t_ms = _safe_float(row.get("t_video_ms")) or _timecode_to_ms(row.get("video_timecode"))
        if t_ms is not None:
            times.append(t_ms)
    return max(times, default=0.0)


def _draw_activity_timeline(path: Path, analysis: dict[str, Any]) -> None:
    font, small, title_font = _load_fonts()
    width, height = 1100, 620
    image = Image.new("RGB", (width, height), "#f8fafb")
    draw = ImageDraw.Draw(image)
    draw.text((40, 24), "每分钟事件与点击节奏", font=title_font, fill="#17212b")
    left, top, right, bottom = 90, 90, 1040, 540
    _draw_axes(draw, left, top, right, bottom)

    minutes = analysis["minutes"]
    minute_counts = analysis["minute_counts"]
    minute_clicks = analysis["minute_clicks"]
    max_value = max([minute_counts[minute] for minute in minutes] + [1])
    bar_width = (right - left) / max(len(minutes), 1)
    for minute in minutes:
        x0 = left + minute * bar_width + 2
        x1 = left + (minute + 1) * bar_width - 2
        y = bottom - (minute_counts[minute] / max_value) * (bottom - top)
        draw.rectangle((x0, y, x1, bottom), fill="#8ecae6")

    points = [
        (
            left + minute * bar_width + bar_width / 2,
            bottom - (minute_clicks[minute] / max_value) * (bottom - top),
        )
        for minute in minutes
    ]
    if len(points) > 1:
        draw.line(points, fill="#d83b3b", width=4)
        for x, y in points:
            draw.ellipse((x - 4, y - 4, x + 4, y + 4), fill="#d83b3b")

    step = max(1, max_value // 5)
    for value in range(0, max_value + 1, step):
        y = bottom - (value / max_value) * (bottom - top)
        draw.line((left - 5, y, right, y), fill="#e4eaee")
        draw.text((28, y - 9), str(value), font=small, fill="#60717d")
    for minute in range(0, len(minutes), 2):
        x = left + minute * bar_width + bar_width / 2
        draw.text((x - 10, bottom + 12), str(minute), font=small, fill="#60717d")

    draw.text((left, height - 52), "横轴：录制时间（分钟）    纵轴：事件数", font=small, fill="#60717d")
    draw.text((left, height - 30), "蓝色柱：全部事件    红色折线：点击事件", font=small, fill="#60717d")
    image.save(path)


def _draw_click_circle_heatmap(path: Path, analysis: dict[str, Any]) -> None:
    font, small, title_font = _load_fonts()
    region_width = analysis["region_width"]
    region_height = analysis["region_height"]
    clicks = analysis["clicks"]

    max_chart_width = 560
    max_chart_height = 900
    scale = min(max_chart_width / max(1, region_width), max_chart_height / max(1, region_height))
    chart_width = max(1, int(round(region_width * scale)))
    chart_height = max(1, int(round(region_height * scale)))
    width = max(760, chart_width + 160)
    height = max(640, 90 + chart_height + 135)
    image = Image.new("RGB", (width, height), "#f8fafb")
    draw = ImageDraw.Draw(image)
    draw.text((40, 24), "点击圆圈热力图", font=title_font, fill="#17212b")

    left = int((width - chart_width) / 2)
    top = 90
    radius = max(5, int(round(7 * scale)))
    density = [[0.0 for _ in range(chart_width)] for _ in range(chart_height)]
    kernel = _circle_kernel(radius)

    for x, y, _ in clicks:
        px = int(round(x * scale))
        py = int(round(y * scale))
        x0 = max(0, px - radius)
        y0 = max(0, py - radius)
        x1 = min(chart_width, px + radius + 1)
        y1 = min(chart_height, py + radius + 1)
        kernel_x0 = x0 - (px - radius)
        kernel_y0 = y0 - (py - radius)
        for target_y in range(y0, y1):
            density_row = density[target_y]
            kernel_row = kernel[kernel_y0 + target_y - y0]
            for target_x in range(x0, x1):
                density_row[target_x] += kernel_row[kernel_x0 + target_x - x0]

    max_density = max((max(row) for row in density), default=0.0)
    chart = Image.new("RGB", (chart_width, chart_height), "white")
    pixels = chart.load()
    ink = (205, 45, 32)
    for y, row in enumerate(density):
        for x, value in enumerate(row):
            if value <= 0:
                continue
            alpha = min(0.92, 1.0 - math.exp(-value / 3.2))
            pixels[x, y] = tuple(int(255 * (1 - alpha) + channel * alpha) for channel in ink)

    image.paste(chart, (left, top))
    draw.rectangle((left, top, left + chart_width, top + chart_height), outline="#263238", width=2)

    note_y = top + chart_height + 28
    draw.text((40, note_y), f"点击数：{len(clicks)}    录制区域：{region_width}x{region_height}", font=font, fill="#263238")
    draw.text((40, note_y + 30), f"每次点击叠加一个半透明小圆，重叠越多颜色越深；圆半径约 {radius}px", font=small, fill="#60717d")
    draw.text((40, note_y + 54), f"最大叠加密度：{max_density:.1f}", font=small, fill="#60717d")
    image.save(path)


def _circle_kernel(radius: int) -> list[list[float]]:
    kernel: list[list[float]] = []
    for y in range(-radius, radius + 1):
        row: list[float] = []
        for x in range(-radius, radius + 1):
            distance = (x * x + y * y) ** 0.5
            if distance <= radius:
                row.append(1.0 + max(0.0, 1.0 - distance / max(1, radius)) * 0.35)
            else:
                row.append(0.0)
        kernel.append(row)
    return kernel


def _draw_click_scatter(path: Path, analysis: dict[str, Any]) -> None:
    font, small, title_font = _load_fonts()
    region_width = analysis["region_width"]
    region_height = analysis["region_height"]
    clicks = analysis["clicks"]
    width, height = 760, 1060
    image = Image.new("RGB", (width, height), "#f8fafb")
    draw = ImageDraw.Draw(image)
    draw.text((40, 24), "点击位置分布（录制区域）", font=title_font, fill="#17212b")
    left, top = 150, 90
    scale = min(500 / region_width, 860 / region_height)
    chart_width = int(region_width * scale)
    chart_height = int(region_height * scale)
    draw.rectangle((left, top, left + chart_width, top + chart_height), fill="white", outline="#263238", width=2)
    for x, y, _ in clicks:
        px = left + x * scale
        py = top + y * scale
        draw.ellipse((px - 2, py - 2, px + 2, py + 2), fill="#d83b3b")
    draw.text((40, top + chart_height + 28), f"点击总数：{len(clicks)}    录制区域：{region_width}x{region_height}", font=font, fill="#263238")
    draw.text((40, top + chart_height + 58), "每个红点代表一次 click 事件", font=small, fill="#60717d")
    image.save(path)


def _draw_drag_durations(path: Path, analysis: dict[str, Any]) -> None:
    font, small, title_font = _load_fonts()
    metrics = analysis["metrics"]
    values = [item["duration_ms"] for item in analysis["drag_pairs"]] or [0]
    width, height = 980, 560
    image = Image.new("RGB", (width, height), "#f8fafb")
    draw = ImageDraw.Draw(image)
    draw.text((40, 24), "拖拽时长分布", font=title_font, fill="#17212b")
    left, top, right, bottom = 100, 90, 920, 470
    _draw_axes(draw, left, top, right, bottom)
    max_value = max(values) or 1
    bar_width = (right - left) / max(len(values), 1)
    for index, value in enumerate(values):
        x0 = left + index * bar_width + 4
        x1 = left + (index + 1) * bar_width - 4
        y = bottom - (value / max_value) * (bottom - top)
        draw.rectangle((x0, y, x1, bottom), fill="#1f9d55")
        draw.text((x0, bottom + 10), str(index + 1), font=small, fill="#60717d")
    draw.text((left, height - 76), "横轴：拖拽序号    纵轴：时长（毫秒）", font=small, fill="#60717d")
    draw.text(
        (left, height - 52),
        f"平均：{metrics['avg_drag_duration_ms']} ms    中位数：{metrics['median_drag_duration_ms']} ms    最大：{metrics['max_drag_duration_ms']} ms",
        font=font,
        fill="#263238",
    )
    image.save(path)


def _write_behavior_workbook(path: Path, analysis: dict[str, Any]) -> None:
    metrics = analysis["metrics"]
    wb = Workbook()
    ws = wb.active
    ws.title = "概览"
    ws.append(["指标", "数值", "说明"])
    rows = [
        ("录制时长（分钟）", metrics["duration_minutes"], "有效视频时长"),
        ("事件总数", metrics["events_total"], "down/up/click/drag 等事件合计"),
        ("采样总数", metrics["samples_total"], "鼠标位置采样行数"),
        ("实际采样 Hz", metrics["actual_sample_hz"], "接近配置值表示采样稳定"),
        ("点击数", metrics["clicks_total"], "click 事件数"),
        ("有坐标点击数", metrics["clicks_with_position"], "可进入热力图的点击数"),
        ("点击/分钟", metrics["clicks_per_minute"], "操作强度指标"),
        ("双击候选数", metrics["double_click_candidates"], "可能的重复确认点击"),
        ("双击候选占比", metrics["double_click_candidate_ratio"], "double_click_candidate / click"),
        ("滚轮事件数", metrics["wheel_events"], "wheel 事件数"),
        ("拖拽数", metrics["drag_count"], "drag_start/end 配对数或 drag_start 数"),
        ("平均拖拽时长 ms", metrics["avg_drag_duration_ms"], "拖拽复杂度参考"),
        ("底部三分之一点击占比", metrics["bottom_third_click_ratio"], "判断操作是否集中在底栏/按钮区"),
        ("底部中区点击占比", metrics["bottom_center_click_ratio"], "最明显的粗粒度热区"),
        ("录制区域宽", metrics["region_width"], "热力图横向尺寸"),
        ("录制区域高", metrics["region_height"], "热力图纵向尺寸"),
        ("热力图网格", f"{metrics['heatmap_columns']}x{metrics['heatmap_rows']}", "短边约 40 个正方形格，长边按录制比例扩展"),
    ]
    for row in rows:
        ws.append(row)
    _style_header(ws)
    for column, width in [("A", 30), ("B", 18), ("C", 54)]:
        ws.column_dimensions[column].width = width

    ws2 = wb.create_sheet("分钟节奏")
    ws2.append(["分钟", "全部事件", "点击事件", "拖拽事件"])
    for minute in analysis["minutes"]:
        ws2.append(
            [
                minute,
                analysis["minute_counts"][minute],
                analysis["minute_clicks"][minute],
                analysis["minute_drags"][minute],
            ]
        )
    _style_header(ws2)
    line = LineChart()
    line.title = "每分钟事件与点击节奏"
    line.y_axis.title = "事件数"
    line.x_axis.title = "录制时间（分钟）"
    data = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=len(analysis["minutes"]) + 1)
    categories = Reference(ws2, min_col=1, min_row=2, max_row=len(analysis["minutes"]) + 1)
    line.add_data(data, titles_from_data=True)
    line.set_categories(categories)
    line.height = 9
    line.width = 20
    ws2.add_chart(line, "F2")

    ws3 = wb.create_sheet("点击热力图")
    heat_grid = analysis["heat_grid"]
    heatmap_rows = len(heat_grid)
    heatmap_cols = len(heat_grid[0]) if heat_grid else 0
    ws3.append(["网格"] + [f"列{index + 1}" for index in range(heatmap_cols)])
    max_heat = max((max(row) for row in analysis["heat_grid"]), default=0) or 1
    for row_index, values in enumerate(heat_grid, start=1):
        ws3.append([f"行{row_index}", *values])
    _style_header(ws3)
    for row in ws3.iter_rows(
        min_row=2,
        min_col=2,
        max_row=heatmap_rows + 1,
        max_col=heatmap_cols + 1,
    ):
        for cell in row:
            color = _heat_color(cell.value or 0, max_heat)
            cell.fill = PatternFill("solid", fgColor="".join(f"{item:02X}" for item in color))
            cell.alignment = Alignment(horizontal="center")
    ws3.column_dimensions["A"].width = 10
    for index in range(2, heatmap_cols + 2):
        ws3.column_dimensions[ws3.cell(row=1, column=index).column_letter].width = 4

    ws4 = wb.create_sheet("拖拽明细")
    ws4.append(["序号", "开始时间（秒）", "结束时间（秒）", "时长（毫秒）", "距离（px）"])
    for index, item in enumerate(analysis["drag_pairs"], start=1):
        ws4.append(
            [
                index,
                round(item["start_ms"] / 1000, 3),
                round(item["end_ms"] / 1000, 3),
                round(item["duration_ms"], 1),
                round(item["distance_px"], 1),
            ]
        )
    _style_header(ws4)
    if analysis["drag_pairs"]:
        bar = BarChart()
        bar.title = "拖拽时长分布"
        bar.y_axis.title = "时长（毫秒）"
        bar.x_axis.title = "拖拽序号"
        data = Reference(ws4, min_col=4, min_row=1, max_row=len(analysis["drag_pairs"]) + 1)
        categories = Reference(ws4, min_col=1, min_row=2, max_row=len(analysis["drag_pairs"]) + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)
        bar.height = 8
        bar.width = 18
        ws4.add_chart(bar, "G2")

    wb.save(path)


def _style_header(worksheet: Any) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="263238")


def _load_fonts() -> tuple[Any, Any, Any]:
    font_candidates = [
        Path(r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path(r"C:\Windows\Fonts\simsun.ttc"),
        Path(r"C:\Windows\Fonts\arial.ttf"),
    ]
    font_path = next((path for path in font_candidates if path.exists()), None)
    try:
        if font_path is None:
            raise OSError("No local font found")
        return (
            ImageFont.truetype(str(font_path), 18),
            ImageFont.truetype(str(font_path), 13),
            ImageFont.truetype(str(font_path), 24),
        )
    except Exception:
        font = ImageFont.load_default()
        return font, font, font


def _draw_axes(draw: ImageDraw.ImageDraw, left: int, top: int, right: int, bottom: int) -> None:
    draw.line((left, top, left, bottom), fill="#60717d", width=2)
    draw.line((left, bottom, right, bottom), fill="#60717d", width=2)


def _heat_color(value: int | float, max_value: int | float) -> tuple[int, int, int]:
    if max_value <= 0:
        return (235, 240, 243)
    ratio = min(1.0, float(value) / float(max_value))
    if ratio < 0.5:
        t = ratio / 0.5
        return (int(142 + 113 * t), int(202 + 40 * t), int(230 - 150 * t))
    t = (ratio - 0.5) / 0.5
    return (int(255 - 39 * t), int(242 - 159 * t), int(80 - 75 * t))


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _timecode_to_ms(value: Any) -> float | None:
    if not isinstance(value, str):
        return None
    try:
        minutes, rest = value.split(":", 1)
        seconds, millis = rest.split(".", 1)
        return int(minutes) * 60_000 + int(seconds) * 1000 + int(millis[:3].ljust(3, "0"))
    except (ValueError, TypeError):
        return None
