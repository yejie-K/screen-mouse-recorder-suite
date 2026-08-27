from __future__ import annotations

from datetime import datetime
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image

from .charts import (
    render_emotion_timeline_draft,
    render_emotion_timeline_final,
    render_growth_timeline_draft,
    render_growth_timeline_final,
    render_open_timeline,
)
from .handoff import build_open_timeline_agent_spec, render_open_timeline_agent_report, write_text_atomic
from .package import JourneyPackageError, build_semantic_input, write_json_atomic
from .rules import load_rule_file
from .workspace import refresh_journey_workspace, validate_final_gate


def generate_final_product(
    workspace_dir: Path,
    *,
    taxonomy_path: Path,
    emotion_rules_path: Path,
) -> dict[str, Any]:
    workspace_dir = workspace_dir.resolve()
    manifest = validate_final_gate(workspace_dir)
    artifacts = manifest["artifacts"]
    final_dir = workspace_dir / "final"
    existing = [path.name for path in final_dir.iterdir()] if final_dir.exists() else []
    if existing:
        raise JourneyPackageError("正式产物目录必须为空: " + ", ".join(sorted(existing)))
    final_dir.mkdir(parents=True, exist_ok=True)

    confirmed_events_path = _artifact(workspace_dir, artifacts, "confirmed_events")
    confirmed_metrics_path = _artifact(workspace_dir, artifacts, "confirmed_metrics")
    confirmed_events = _read_object(confirmed_events_path)
    confirmed_metrics = _read_object(confirmed_metrics_path)
    events = [
        item for item in confirmed_events.get("events") or []
        if (item.get("semantic_review") or {}).get("status") == "confirmed"
    ]
    metrics = [
        item for item in confirmed_metrics.get("metrics") or []
        if (item.get("review") or {}).get("status") == "confirmed"
    ]
    if not events:
        raise JourneyPackageError("没有人工确认事件，无法生成正式产物")
    evidence_dir = final_dir / "evidence"
    chart_dir = final_dir / "charts"
    evidence_dir.mkdir()
    chart_dir.mkdir()

    taxonomy = load_rule_file(taxonomy_path)
    emotion_rules = load_rule_file(emotion_rules_path)
    semantic_input = build_semantic_input(
        confirmed_events_path,
        confirmed_events,
        taxonomy,
        emotion_rules,
        session_id=str(manifest["session"]["session_id"]),
        total_play_time_ms=int(manifest["session"]["duration_ms"]),
    )
    evidence = _materialize_event_tiles(workspace_dir, events, evidence_dir)
    workbook_path = final_dir / "游戏历程拆解结果.xlsx"
    _write_workbook(workbook_path, manifest, events, metrics, evidence)

    open_path = chart_dir / "玩法系统开放节奏.png"
    emotion_path = chart_dir / "事件情绪时间图.png"
    growth_path = chart_dir / "成长反馈时间图.png"
    open_report = render_open_timeline(semantic_input, open_path, game_name=str(manifest["session"]["game_name"]))
    emotion_report = render_emotion_timeline_final(confirmed_events, emotion_path)
    growth_report = render_growth_timeline_final(confirmed_metrics, growth_path)
    write_json_atomic(chart_dir / "玩法系统开放节奏.report.json", open_report)
    write_json_atomic(chart_dir / "事件情绪时间图.report.json", emotion_report)
    write_json_atomic(chart_dir / "成长反馈时间图.report.json", growth_report)
    open_handoff = _write_open_timeline_handoff(
        chart_dir,
        semantic_input,
        open_report,
        taxonomy,
        game_name=str(manifest["session"]["game_name"]),
        base_name="玩法系统开放节奏",
        candidate_mode=False,
    )

    report = {
        "schema_version": "1.0",
        "task_id": "JOURNEY_FINAL_PRODUCT_V1",
        "status": "complete",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "session": manifest["session"],
        "inputs": {
            "confirmed_events": {"sha256": _sha256_file(confirmed_events_path), "confirmed_count": len(events)},
            "confirmed_metrics": {"sha256": _sha256_file(confirmed_metrics_path), "confirmed_count": len(metrics)},
        },
        "outputs": {
            "workbook": workbook_path.name,
            "charts": [
                f"charts/{open_path.name}",
                f"charts/{emotion_path.name}",
                f"charts/{growth_path.name}",
            ],
            "event_evidence": [f"evidence/{path.name}" for path in evidence.values()],
            "open_timeline_handoff": [f"charts/{name}" for name in open_handoff],
        },
    }
    write_json_atomic(final_dir / "manifest.json", report)
    (final_dir / "AGENT_README.md").write_text(_agent_report(report), encoding="utf-8")
    workspace_manifest = refresh_journey_workspace(workspace_dir)
    write_json_atomic(workspace_dir / "journey_workspace.json", workspace_manifest)
    return report


def generate_preview_product(
    workspace_dir: Path,
) -> dict[str, Any]:
    workspace_dir = workspace_dir.resolve()
    manifest = refresh_journey_workspace(workspace_dir)
    artifacts = manifest["artifacts"]
    event_dir = _artifact(workspace_dir, artifacts, "event_review_manifest").parent
    semantic_input_path = event_dir / "journey_semantic_input.json"
    semantic_output_path = event_dir / "journey_semantic_output.json"
    metric_candidates_path = _artifact(workspace_dir, artifacts, "metric_observations")
    semantic_input = _read_object(semantic_input_path)
    semantic_output = _read_object(semantic_output_path)
    metric_candidates = _read_object(metric_candidates_path)
    if str((semantic_input.get("session") or {}).get("session_id") or "") != str(manifest["session"]["session_id"]):
        raise JourneyPackageError("语义候选Session与工作空间不一致")
    if str((metric_candidates.get("session") or {}).get("session_id") or "") != str(manifest["session"]["session_id"]):
        raise JourneyPackageError("指标候选Session与工作空间不一致")

    preview_dir = workspace_dir / "preview"
    if preview_dir.exists() and any(preview_dir.iterdir()):
        raise JourneyPackageError("预览产物目录必须为空")
    preview_dir.mkdir(parents=True, exist_ok=True)
    evidence_dir = preview_dir / "evidence"
    chart_dir = preview_dir / "charts"
    evidence_dir.mkdir()
    chart_dir.mkdir()
    events = list(semantic_input.get("events") or [])
    annotations = {
        str(item.get("event_id") or ""): item
        for item in semantic_output.get("event_annotations") or []
        if isinstance(item, dict)
    }
    evidence = _materialize_event_tiles(workspace_dir, events, evidence_dir)
    workbook_path = preview_dir / "游戏历程拆解候选预览.xlsx"
    _write_preview_workbook(
        workbook_path,
        manifest,
        events,
        annotations,
        list(metric_candidates.get("metrics") or []),
        evidence,
    )
    open_path = chart_dir / "玩法系统开放节奏_候选预览.png"
    emotion_path = chart_dir / "事件情绪时间图_候选预览.png"
    growth_path = chart_dir / "成长反馈时间图_候选预览.png"
    open_report = render_open_timeline(
        semantic_input,
        open_path,
        game_name=f"{manifest['session']['game_name']}（候选预览）",
        candidate_mode=True,
    )
    repository_root = Path(__file__).resolve().parents[3]
    taxonomy = load_rule_file(repository_root / "rules" / "gameplay_taxonomy_v0.1.json")
    emotion_rules = load_rule_file(repository_root / "rules" / "emotion_rules_v0.1.json")
    emotion_report = render_emotion_timeline_draft(semantic_input, emotion_rules, emotion_path)
    growth_report = render_growth_timeline_draft(metric_candidates, growth_path)
    for name, report_payload in (
        ("玩法系统开放节奏_候选预览.report.json", open_report),
        ("事件情绪时间图_候选预览.report.json", emotion_report),
        ("成长反馈时间图_候选预览.report.json", growth_report),
    ):
        write_json_atomic(chart_dir / name, report_payload)
    open_handoff = _write_open_timeline_handoff(
        chart_dir,
        semantic_input,
        open_report,
        taxonomy,
        game_name=str(manifest["session"]["game_name"]),
        base_name="玩法系统开放节奏_候选预览",
        candidate_mode=True,
    )
    report = {
        "schema_version": "1.0",
        "task_id": "JOURNEY_PREVIEW_PRODUCT_V1",
        "status": "draft",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "session": manifest["session"],
        "warning": "包含待复核OCR/语义候选，仅用于推进产物链路，不可作为正式结论",
        "inputs": {
            "event_candidates": {"sha256": _sha256_file(semantic_input_path), "count": len(events)},
            "metric_candidates": {"sha256": _sha256_file(metric_candidates_path), "count": len(metric_candidates.get("metrics") or [])},
        },
        "outputs": {
            "workbook": workbook_path.name,
            "charts": [f"charts/{open_path.name}", f"charts/{emotion_path.name}", f"charts/{growth_path.name}"],
            "event_evidence": [f"evidence/{path.name}" for path in evidence.values()],
            "open_timeline_handoff": [f"charts/{name}" for name in open_handoff],
        },
    }
    write_json_atomic(preview_dir / "manifest.json", report)
    (preview_dir / "AGENT_README.md").write_text(_preview_agent_report(report), encoding="utf-8")
    workspace_manifest = refresh_journey_workspace(workspace_dir)
    write_json_atomic(workspace_dir / "journey_workspace.json", workspace_manifest)
    return report


def _write_open_timeline_handoff(
    chart_dir: Path,
    semantic_input: dict[str, Any],
    chart_report: dict[str, Any],
    taxonomy: dict[str, Any],
    *,
    game_name: str,
    base_name: str,
    candidate_mode: bool,
) -> list[str]:
    input_name = f"{base_name}.input.json"
    taxonomy_name = f"{base_name}.taxonomy.json"
    spec_name = f"{base_name}.agent_spec.json"
    report_name = f"{base_name}.agent_report.md"
    contract_name = f"{base_name}.contract.md"
    render_report_name = f"{base_name}.report.json"
    write_json_atomic(chart_dir / input_name, semantic_input)
    write_json_atomic(chart_dir / taxonomy_name, taxonomy)
    spec = build_open_timeline_agent_spec(
        semantic_input,
        chart_report,
        taxonomy,
        game_name=game_name,
        candidate_mode=candidate_mode,
        source_files={
            "semantic_input": input_name,
            "gameplay_taxonomy": taxonomy_name,
            "chart_contract": contract_name,
            "render_report": render_report_name,
            "rendered_image": chart_report["output"],
        },
    )
    write_json_atomic(chart_dir / spec_name, spec)
    write_text_atomic(chart_dir / report_name, render_open_timeline_agent_report(spec))
    contract_source = Path(__file__).resolve().parents[3] / "docs" / "gameplay_open_timeline_contract.md"
    write_text_atomic(chart_dir / contract_name, contract_source.read_text(encoding="utf-8"))
    return [input_name, taxonomy_name, spec_name, report_name, contract_name]


def _write_workbook(
    target: Path,
    manifest: dict[str, Any],
    events: list[dict[str, Any]],
    metrics: list[dict[str, Any]],
    evidence: dict[str, Path],
) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "说明"
    overview.append(["游戏历程拆解结果"])
    overview.append(["游戏", manifest["session"]["game_name"]])
    overview.append(["Session", manifest["session"]["session_id"]])
    overview.append(["时长(ms)", manifest["session"]["duration_ms"]])
    overview.append(["确认事件", len(events)])
    overview.append(["确认指标", len(metrics)])
    overview.column_dimensions["A"].width = 18
    overview.column_dimensions["B"].width = 42
    overview["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="263238")

    event_sheet = workbook.create_sheet("事件单")
    event_headers = ["时间", "事件名称", "模式", "事件标签", "玩法分类", "情绪分值", "OCR文本", "证据ID", "截图"]
    event_sheet.append(event_headers)
    _style_header(event_sheet, len(event_headers))
    for row_index, event in enumerate(events, start=2):
        semantic = event.get("semantic") or {}
        event_sheet.append([
            event.get("timestamp") or "",
            event.get("event_name") or "",
            semantic.get("mode_tag") or "",
            semantic.get("event_tag") or "",
            semantic.get("event_category") or "",
            semantic.get("emotion_score"),
            event.get("ocr_excerpt") or "",
            event.get("event_id") or "",
            "",
        ])
        event_sheet.row_dimensions[row_index].height = 150
        image_path = evidence.get(str(event.get("event_id") or ""))
        if image_path is not None:
            image = WorksheetImage(str(image_path))
            image.width = 112
            image.height = 210
            event_sheet.add_image(image, f"I{row_index}")
    widths = [16, 24, 12, 18, 18, 12, 70, 28, 18]
    for index, width in enumerate(widths, start=1):
        event_sheet.column_dimensions[_column_letter(index)].width = width
    event_sheet.freeze_panes = "A2"
    _style_body(event_sheet, 2, event_sheet.max_row, len(event_headers))

    metric_sheet = workbook.create_sheet("指标变化")
    metric_headers = ["时间", "指标", "识别原文", "确认值", "解析字段", "区域", "证据ID"]
    metric_sheet.append(metric_headers)
    _style_header(metric_sheet, len(metric_headers))
    for metric in metrics:
        metric_sheet.append([
            metric.get("timestamp") or "",
            metric.get("metric_key") or "",
            metric.get("raw_text") or "",
            metric.get("parsed_value"),
            json.dumps(metric.get("parsed_fields") or {}, ensure_ascii=False),
            metric.get("region_id") or "",
            metric.get("observation_id") or "",
        ])
    widths = [16, 18, 40, 18, 26, 28, 32]
    for index, width in enumerate(widths, start=1):
        metric_sheet.column_dimensions[_column_letter(index)].width = width
    metric_sheet.freeze_panes = "A2"
    _style_body(metric_sheet, 2, metric_sheet.max_row, len(metric_headers))
    workbook.save(target)


def _write_preview_workbook(
    target: Path,
    manifest: dict[str, Any],
    events: list[dict[str, Any]],
    annotations: dict[str, dict[str, Any]],
    metrics: list[dict[str, Any]],
    evidence: dict[str, Path],
) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "预览说明"
    rows = [
        ["游戏历程拆解候选预览"],
        ["状态", "草稿 / 待复核"],
        ["用途", "继续验证XLSX和图表链路，不作为正式分析结论"],
        ["游戏", manifest["session"]["game_name"]],
        ["Session", manifest["session"]["session_id"]],
        ["事件候选", len(events)],
        ["指标候选", len(metrics)],
    ]
    for row in rows:
        overview.append(row)
    overview.column_dimensions["A"].width = 20
    overview.column_dimensions["B"].width = 70
    overview["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="A44848")

    event_sheet = workbook.create_sheet("事件候选")
    event_headers = ["时间", "事件名称", "模式候选", "事件标签候选", "情绪候选", "状态", "OCR文本", "证据ID", "截图"]
    event_sheet.append(event_headers)
    _style_header(event_sheet, len(event_headers), color="A44848")
    for row_index, event in enumerate(events, start=2):
        event_id = str(event.get("event_id") or "")
        annotation = annotations.get(event_id) or {}
        event_sheet.append([
            event.get("timestamp") or "",
            event.get("event_name") or "",
            annotation.get("mode_tag") or "待判断",
            annotation.get("event_tag") or "其他开放",
            annotation.get("emotion_score_candidate"),
            "待复核",
            event.get("ocr_excerpt") or "",
            event_id,
            "",
        ])
        event_sheet.row_dimensions[row_index].height = 150
        image_path = evidence.get(event_id)
        if image_path is not None:
            image = WorksheetImage(str(image_path))
            image.width = 112
            image.height = 210
            event_sheet.add_image(image, f"I{row_index}")
    for index, width in enumerate([16, 24, 14, 20, 12, 12, 70, 30, 18], start=1):
        event_sheet.column_dimensions[_column_letter(index)].width = width
    event_sheet.freeze_panes = "A2"
    _style_body(event_sheet, 2, event_sheet.max_row, len(event_headers))

    metric_sheet = workbook.create_sheet("指标候选")
    metric_headers = ["时间", "指标类型", "OCR原文", "解析值", "置信度", "状态", "区域", "证据图路径", "证据ID"]
    metric_sheet.append(metric_headers)
    _style_header(metric_sheet, len(metric_headers), color="A44848")
    for metric in metrics:
        crops = list((metric.get("evidence") or {}).get("crop_images") or [])
        metric_sheet.append([
            metric.get("timestamp") or "",
            metric.get("metric_key") or "",
            metric.get("raw_text") or "",
            metric.get("parsed_value"),
            metric.get("confidence"),
            "待复核",
            metric.get("region_id") or "",
            crops[0] if crops else "",
            metric.get("observation_id") or "",
        ])
    for index, width in enumerate([16, 18, 42, 18, 12, 12, 28, 52, 34], start=1):
        metric_sheet.column_dimensions[_column_letter(index)].width = width
    metric_sheet.freeze_panes = "A2"
    _style_body(metric_sheet, 2, metric_sheet.max_row, len(metric_headers))
    workbook.save(target)


def _materialize_event_tiles(workspace_dir: Path, events: list[dict[str, Any]], target_dir: Path) -> dict[str, Path]:
    runtime = _read_object(workspace_dir / "runtime" / "review_session.json")
    sheet_meta = {str(item.get("name") or ""): item for item in runtime.get("contactSheets") or []}
    result: dict[str, Path] = {}
    for event in events:
        event_id = str(event.get("event_id") or "")
        evidence = event.get("evidence") or {}
        sheet_name = str(evidence.get("contact_sheet") or "")
        meta = sheet_meta.get(sheet_name)
        row = evidence.get("sheet_row")
        column = evidence.get("sheet_col")
        source = workspace_dir / "runtime" / "contact_sheets" / sheet_name
        if not meta or not source.is_file() or not isinstance(row, int) or not isinstance(column, int):
            continue
        with Image.open(source) as sheet:
            rows = max(1, int(meta.get("rows") or 1))
            columns = max(1, int(meta.get("columns") or 1))
            gap, header = 8, 46
            thumb_width = (sheet.width - (columns + 1) * gap) // columns
            label_height = max(22, min(34, thumb_width // 11))
            cell_height = (sheet.height - header - (rows + 1) * gap) // rows
            thumb_height = max(1, cell_height - label_height)
            left = gap + (column - 1) * (thumb_width + gap)
            top = header + gap + (row - 1) * (cell_height + gap)
            tile = sheet.crop((left, top, left + thumb_width, top + thumb_height)).convert("RGB")
            target = target_dir / f"{event_id}.jpg"
            tile.save(target, "JPEG", quality=90, optimize=True)
            tile.close()
            result[event_id] = target
    return result


def _style_header(sheet, columns: int, *, color: str = "46689A") -> None:
    fill = PatternFill("solid", fgColor=color)
    border = Border(bottom=Side(style="thin", color="C9D0D7"))
    for cell in sheet[1][:columns]:
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.row_dimensions[1].height = 30


def _style_body(sheet, start: int, end: int, columns: int) -> None:
    border = Border(
        left=Side(style="thin", color="D9DEE5"), right=Side(style="thin", color="D9DEE5"),
        top=Side(style="thin", color="D9DEE5"), bottom=Side(style="thin", color="D9DEE5"),
    )
    for row in sheet.iter_rows(min_row=start, max_row=max(start, end), max_col=columns):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10, color="24313D")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def _column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _artifact(root: Path, artifacts: dict[str, Any], key: str) -> Path:
    result = (root / str(artifacts[key])).resolve()
    result.relative_to(root)
    return result


def _read_object(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, dict):
        raise JourneyPackageError(f"JSON顶层必须是对象: {path}")
    return payload


def _sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _agent_report(report: dict[str, Any]) -> str:
    session = report["session"]
    return (
        "# 游戏历程拆解正式产物\n\n"
        f"- 游戏：{session['game_name']}\n"
        f"- Session：{session['session_id']}\n"
        f"- 确认事件：{report['inputs']['confirmed_events']['confirmed_count']}\n"
        f"- 确认指标：{report['inputs']['confirmed_metrics']['confirmed_count']}\n\n"
        "本目录只使用人工确认后的功能事件和指标。`manifest.json`记录输入指纹与全部输出，"
        "XLSX中的证据ID可回溯到工作空间复核文件，三张PNG由确定性脚本生成。\n"
    )


def _preview_agent_report(report: dict[str, Any]) -> str:
    session = report["session"]
    return (
        "# 游戏历程拆解候选预览\n\n"
        f"- 游戏：{session['game_name']}\n"
        f"- Session：{session['session_id']}\n"
        f"- 事件候选：{report['inputs']['event_candidates']['count']}\n"
        f"- 指标候选：{report['inputs']['metric_candidates']['count']}\n\n"
        "> 本目录包含未人工复核的OCR和语义候选，只用于验证后续XLSX与图表链路，不能作为正式分析结论。\n"
    )
